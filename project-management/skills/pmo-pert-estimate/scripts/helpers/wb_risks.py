"""Risks sheet builder for pmo-pert-estimate Excel output.

Generates a 'Risks' worksheet containing risk register data with
P×I scoring, priority classification, and contingency calculations.
"""
from openpyxl import Workbook
from openpyxl.styles import Font

from helpers.formatting import (
    HEADER_FONT,
    apply_column_widths,
    apply_style,
    get_total_style,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_SHEET_NAME = "Risks"
_DATA_START_ROW = 2

_RED_FONT = Font(bold=True, color="FFFF0000")
_HIGH_SCORE_THRESHOLD = 15

_COLUMN_WIDTHS = {
    "A": 8,
    "B": 35,
    "C": 18,
    "D": 20,
    "E": 14,
    "F": 12,
    "G": 12,
    "H": 12,
    "I": 14,
    "J": 35,
    "K": 10,
    "L": 16,
    "M": 18,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build(wb: Workbook, data: dict, wbs_info: dict | None = None) -> dict:
    """Build the Risks sheet and append it to the workbook.

    Args:
        wb:       An openpyxl Workbook instance.
        data:     Input dict with keys ``config``, ``risks``, ``roles``, ``phases``.
        wbs_info: Optional WBS sheet_info. When provided, the Management
                  Reserve formula uses the correct PMI-compliant base
                  (Tech + Overhead + Contingency) cross-referenced from the
                  WBS total row. When omitted, the legacy
                  ``=L{total}*pct`` formula is preserved for backward
                  compatibility with callers/tests that don't track WBS.

    Returns:
        A ``sheet_info`` dict with keys:
        - ``sheet_name``
        - ``data_start_row``
        - ``data_end_row``
        - ``total_contingency_row``
        - ``reserve_row``
    """
    config = data["config"]
    risks = data["risks"]

    effort_unit = config.get("effort_unit", "pd")
    avg_rate = config.get("avg_rate")
    reserve_pct = config.get("management_reserve_pct", 0.10)
    pm_pct = float(config.get("pm_overhead_pct") or 0)
    devops_pct = float(config.get("devops_overhead_pct") or 0)

    ws = wb.create_sheet(title=_SHEET_NAME)

    # -----------------------------------------------------------------------
    # Row 1: Headers
    # -----------------------------------------------------------------------
    headers = [
        "ID",
        "Risk Description",
        "Category",
        "Affected Phases",
        "Probability (1-5)",
        "Impact (1-5)",
        "Risk Score",
        "Priority",
        "Strategy",
        "Mitigation Action",
        "Owner",
        f"Contingency ({effort_unit})",
        "Contingency Cost",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, size=11)

    # -----------------------------------------------------------------------
    # Data rows
    # -----------------------------------------------------------------------
    data_start = _DATA_START_ROW
    data_end = data_start + len(risks) - 1

    for i, risk in enumerate(risks):
        row_num = data_start + i
        score = risk["probability"] * risk["impact"]

        # Column A: ID
        ws.cell(row=row_num, column=1, value=risk["id"])
        # Column B: Description
        ws.cell(row=row_num, column=2, value=risk["description"])
        # Column C: Category
        ws.cell(row=row_num, column=3, value=risk["category"])
        # Column D: Affected Phases (comma-joined)
        affected = ", ".join(risk.get("affected_phases", []))
        ws.cell(row=row_num, column=4, value=affected)
        # Column E: Probability (numeric)
        ws.cell(row=row_num, column=5, value=risk["probability"])
        # Column F: Impact (numeric)
        ws.cell(row=row_num, column=6, value=risk["impact"])
        # Column G: Risk Score formula =E{n}*F{n}
        ws.cell(row=row_num, column=7, value=f"=E{row_num}*F{row_num}")
        # Column H: Priority nested IF formula
        g_ref = f"G{row_num}"
        priority_formula = (
            f'=IF({g_ref}>=15,"CRITICAL",'
            f'IF({g_ref}>=10,"HIGH",'
            f'IF({g_ref}>=5,"MEDIUM","LOW")))'
        )
        ws.cell(row=row_num, column=8, value=priority_formula)
        # Column I: Strategy
        ws.cell(row=row_num, column=9, value=risk.get("strategy", ""))
        # Column J: Mitigation Action
        ws.cell(row=row_num, column=10, value=risk.get("mitigation", ""))
        # Column K: Owner
        ws.cell(row=row_num, column=11, value=risk.get("owner", ""))
        # Column L: Contingency effort
        ws.cell(row=row_num, column=12, value=risk.get("contingency_effort"))
        # Column M: Contingency cost formula (only when avg_rate is set)
        if avg_rate is not None:
            ws.cell(row=row_num, column=13, value=f"=L{row_num}*{avg_rate}")

        # Apply red font to high-score risk rows (score >= threshold)
        if score >= _HIGH_SCORE_THRESHOLD:
            for col_idx in range(1, 14):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.font = Font(bold=True, color="FFFF0000")

    # -----------------------------------------------------------------------
    # Footer rows: blank separator, TOTAL, Reserve
    # -----------------------------------------------------------------------
    blank_row = data_end + 1
    total_row = blank_row + 1
    reserve_row = total_row + 1

    # TOTAL row
    total_cell_a = ws.cell(row=total_row, column=1, value="TOTAL")
    apply_style(total_cell_a, get_total_style())

    total_l = ws.cell(
        row=total_row,
        column=12,
        value=f"=SUM(L{data_start}:L{data_end})",
    )
    apply_style(total_l, get_total_style())

    if avg_rate is not None:
        total_m = ws.cell(
            row=total_row,
            column=13,
            value=f"=SUM(M{data_start}:M{data_end})",
        )
        apply_style(total_m, get_total_style())

    # Management Reserve row.
    # When wbs_info is supplied, MR is computed on the correct PMI base:
    #   Tech PERT (WBS!H{total}) × (1 + pm_pct + devops_pct) + Contingency
    # — i.e. the same "Fascia BASSA" value the Summary sheet shows.
    # Otherwise we fall back to the legacy formula for callers that did not
    # build a WBS sheet first.
    ws.cell(row=reserve_row, column=1, value="Management Reserve")
    if wbs_info is not None:
        wbs_total_row = wbs_info["total_row"]
        mr_formula = (
            f"=(WBS!H{wbs_total_row}*(1+{pm_pct}+{devops_pct})"
            f"+L{total_row})*{reserve_pct}"
        )
    else:
        mr_formula = f"=L{total_row}*{reserve_pct}"
    ws.cell(row=reserve_row, column=12, value=mr_formula)
    if avg_rate is not None:
        # Cost projection mirrors the effort formula scaled by avg_rate.
        if wbs_info is not None:
            wbs_total_row = wbs_info["total_row"]
            mr_cost_formula = (
                f"=(WBS!H{wbs_total_row}*(1+{pm_pct}+{devops_pct})"
                f"+L{total_row})*{reserve_pct}*{avg_rate}"
            )
        else:
            mr_cost_formula = f"=M{total_row}*{reserve_pct}"
        ws.cell(row=reserve_row, column=13, value=mr_cost_formula)

    # -----------------------------------------------------------------------
    # Column widths
    # -----------------------------------------------------------------------
    apply_column_widths(ws, _COLUMN_WIDTHS)

    return {
        "sheet_name": _SHEET_NAME,
        "data_start_row": data_start,
        "data_end_row": data_end,
        "total_contingency_row": total_row,
        "reserve_row": reserve_row,
    }
