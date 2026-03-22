"""WBS sheet generator for pmo-pert-estimate Excel output.

Writes a "WBS" sheet into the given workbook with one row per phase,
work package, and leaf activity, applying PERT formulas and styling.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from helpers.formatting import (
    apply_column_widths,
    apply_style,
    get_formula_fill,
    get_leaf_style,
    get_phase_style,
    get_total_style,
    get_wp_style,
    HEADER_FONT,
    NUMBER_FMT,
)


# ---------------------------------------------------------------------------
# Column constants (1-based indices and letters)
# ---------------------------------------------------------------------------

COL_ID = 1          # A
COL_PHASE = 2       # B
COL_WP = 3          # C
COL_ACTIVITY = 4    # D
COL_BEST_E = 5      # E  — best effort
COL_LIKELY_E = 6    # F  — likely effort
COL_WORST_E = 7     # G  — worst effort
COL_PERT_E = 8      # H  — PERT effort (formula)
COL_BEST_D = 9      # I  — best duration
COL_LIKELY_D = 10   # J  — likely duration
COL_WORST_D = 11    # K  — worst duration
COL_PERT_D = 12     # L  — PERT duration (formula)
COL_SIGMA = 13      # M  — σ duration (formula)
COL_RESOURCES = 14  # N
COL_DEPS = 15       # O
COL_RISKS = 16      # P
COL_NOTES = 17      # Q
COL_BILLABLE = 18   # R
COL_BILL_E = 19     # S  — billable PERT effort (formula)

# Columns that hold formula results and should get the formula fill colour
FORMULA_COLS = ("H", "L", "M", "S")

# Column widths (letters → widths in characters)
COLUMN_WIDTHS = {
    "A": 8,
    "B": 30,
    "C": 30,
    "D": 35,
    "E": 14,
    "F": 14,
    "G": 14,
    "H": 14,
    "I": 14,
    "J": 14,
    "K": 14,
    "L": 14,
    "M": 12,
    "N": 25,
    "O": 20,
    "P": 20,
    "Q": 25,
    "R": 10,
    "S": 18,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build(wb: Workbook, data: dict) -> dict:
    """Build the WBS sheet and insert it into *wb*.

    Args:
        wb:   An :class:`openpyxl.Workbook` instance.
        data: Input dict with keys ``config``, ``phases``, ``roles``.

    Returns:
        A ``sheet_info`` dict::

            {
                "sheet_name":    "WBS",
                "data_start_row": 2,
                "data_end_row":   N,
                "phase_rows":    [row_numbers, ...],
                "total_row":     N + 1,
            }
    """
    config = data["config"]
    phases = data["phases"]

    eu = config.get("effort_unit", "pd")
    du = config.get("duration_unit", "d")
    primary_color = config.get("primary_color", "1B4FA5")

    ws: Worksheet = wb.create_sheet("WBS")

    # --- Header row (row 1) ---
    _write_header(ws, eu, du)

    # --- Data rows ---
    current_row = 2
    phase_rows: list[int] = []

    for phase in phases:
        phase_row = current_row
        phase_rows.append(phase_row)
        current_row += 1  # reserve; fill formulas after children written

        wp_info_list: list[dict] = []

        for wp in phase.get("work_packages", []):
            wp_row = current_row
            current_row += 1  # reserve WP row

            first_leaf = current_row
            leaf_rows_for_wp: list[int] = []

            for activity in wp.get("activities", []):
                _write_leaf_row(ws, current_row, activity)
                leaf_rows_for_wp.append(current_row)
                current_row += 1

            last_leaf = current_row - 1

            # Fill WP row now that we know child range
            _write_wp_row(ws, wp_row, wp, first_leaf, last_leaf)
            wp_info_list.append({
                "row": wp_row,
                "first_leaf": first_leaf,
                "last_leaf": last_leaf,
                "leaf_rows": leaf_rows_for_wp,
            })

        # Collect all leaf rows belonging to this phase
        all_leaf_rows = [r for wi in wp_info_list for r in wi["leaf_rows"]]

        # Fill phase row with formulas over ALL leaf rows of this phase
        _write_phase_row(ws, phase_row, phase, all_leaf_rows)

    data_end_row = current_row - 1
    total_row = current_row

    # --- TOTAL row ---
    _write_total_row(ws, total_row, phase_rows)

    # --- Styling ---
    formula_fill = get_formula_fill()
    phase_style = get_phase_style(primary_color)
    wp_style = get_wp_style()
    leaf_style = get_leaf_style()
    total_style = get_total_style()

    for r in range(2, total_row + 1):
        if r in phase_rows:
            row_style = phase_style
        elif r == total_row:
            row_style = total_style
        else:
            row_id = ws[f"A{r}"].value
            if row_id and isinstance(row_id, str) and row_id.count(".") == 1:
                row_style = wp_style
            else:
                row_style = leaf_style

        for col_idx in range(1, 20):
            cell = ws.cell(row=r, column=col_idx)
            apply_style(cell, row_style)

        # Override formula columns with formula fill
        for col_letter in FORMULA_COLS:
            cell = ws[f"{col_letter}{r}"]
            cell.fill = formula_fill

    # --- Number format for numeric / formula columns ---
    for r in range(2, total_row + 1):
        for col_letter in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "S"):
            ws[f"{col_letter}{r}"].number_format = NUMBER_FMT

    # --- Column widths ---
    apply_column_widths(ws, COLUMN_WIDTHS)

    return {
        "sheet_name": "WBS",
        "data_start_row": 2,
        "data_end_row": data_end_row,
        "phase_rows": phase_rows,
        "total_row": total_row,
    }


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _write_header(ws: Worksheet, effort_unit: str, duration_unit: str) -> None:
    headers = [
        "ID",
        "Phase",
        "Work Package",
        "Activity",
        f"Best Effort ({effort_unit})",
        f"Likely Effort ({effort_unit})",
        f"Worst Effort ({effort_unit})",
        f"PERT Effort ({effort_unit})",
        f"Best Duration ({duration_unit})",
        f"Likely Duration ({duration_unit})",
        f"Worst Duration ({duration_unit})",
        f"PERT Duration ({duration_unit})",
        "\u03c3 Duration",
        "Resources",
        "Dependencies",
        "Risks",
        "Notes",
        "Billable",
        "Billable PERT Effort",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT


def _pert_effort_formula(r: int) -> str:
    return f"=(E{r}+4*F{r}+G{r})/6"


def _pert_duration_formula(r: int) -> str:
    return f"=(I{r}+4*J{r}+K{r})/6"


def _sigma_formula(r: int) -> str:
    return f"=(K{r}-I{r})/6"


def _billable_formula(r: int) -> str:
    return f'=IF(R{r}="Y",H{r},0)'


def _write_leaf_row(ws: Worksheet, r: int, activity: dict) -> None:
    ws.cell(row=r, column=COL_ID, value=activity["id"])
    ws.cell(row=r, column=COL_ACTIVITY, value=activity["name"])

    ws.cell(row=r, column=COL_BEST_E, value=activity["best_effort"])
    ws.cell(row=r, column=COL_LIKELY_E, value=activity["likely_effort"])
    ws.cell(row=r, column=COL_WORST_E, value=activity["worst_effort"])
    ws.cell(row=r, column=COL_PERT_E, value=_pert_effort_formula(r))

    ws.cell(row=r, column=COL_BEST_D, value=activity["best_duration"])
    ws.cell(row=r, column=COL_LIKELY_D, value=activity["likely_duration"])
    ws.cell(row=r, column=COL_WORST_D, value=activity["worst_duration"])
    ws.cell(row=r, column=COL_PERT_D, value=_pert_duration_formula(r))

    ws.cell(row=r, column=COL_SIGMA, value=_sigma_formula(r))

    resources = activity.get("resources", [])
    ws.cell(row=r, column=COL_RESOURCES, value=", ".join(resources))

    deps = activity.get("dependencies", [])
    ws.cell(row=r, column=COL_DEPS, value=", ".join(deps))

    risks = activity.get("risks", [])
    ws.cell(row=r, column=COL_RISKS, value=", ".join(risks))

    ws.cell(row=r, column=COL_NOTES, value=activity.get("notes", ""))

    billable_flag = "Y" if activity.get("billable", True) else "N"
    ws.cell(row=r, column=COL_BILLABLE, value=billable_flag)
    ws.cell(row=r, column=COL_BILL_E, value=_billable_formula(r))


def _sum_of_rows(col: str, row_list: list[int]) -> str:
    """Return =SUM(C3,C5,...) formula referencing specific rows."""
    refs = ",".join(f"{col}{r}" for r in row_list)
    return f"=SUM({refs})"


def _sum_range(col: str, first: int, last: int) -> str:
    """Return =SUM(C3:C7) formula for a contiguous range."""
    return f"=SUM({col}{first}:{col}{last})"


def _write_wp_row(ws: Worksheet, r: int, wp: dict,
                  first_leaf: int, last_leaf: int) -> None:
    ws.cell(row=r, column=COL_ID, value=wp["id"])
    ws.cell(row=r, column=COL_WP, value=wp["name"])

    leaf_rows = list(range(first_leaf, last_leaf + 1))

    for col in ("E", "F", "G"):
        ws.cell(row=r, column=_col_idx(col),
                value=_sum_range(col, first_leaf, last_leaf))

    # PERT effort on WP row (aggregated from leaf H column)
    ws.cell(row=r, column=COL_PERT_E, value=_pert_effort_formula(r))

    # Duration aggregates from leaf rows
    for col in ("I", "J", "K"):
        ws.cell(row=r, column=_col_idx(col),
                value=_sum_range(col, first_leaf, last_leaf))

    ws.cell(row=r, column=COL_PERT_D, value=_pert_duration_formula(r))
    ws.cell(row=r, column=COL_SIGMA, value=_sigma_formula(r))
    ws.cell(row=r, column=COL_BILL_E, value=_sum_range("S", first_leaf, last_leaf))


def _write_phase_row(ws: Worksheet, r: int, phase: dict,
                     all_leaf_rows: list[int]) -> None:
    ws.cell(row=r, column=COL_ID, value=phase["id"])
    ws.cell(row=r, column=COL_PHASE, value=phase["name"])

    for col in ("E", "F", "G"):
        ws.cell(row=r, column=_col_idx(col),
                value=_sum_of_rows(col, all_leaf_rows))

    ws.cell(row=r, column=COL_PERT_E, value=_pert_effort_formula(r))

    for col in ("I", "J", "K"):
        ws.cell(row=r, column=_col_idx(col),
                value=_sum_of_rows(col, all_leaf_rows))

    ws.cell(row=r, column=COL_PERT_D, value=_pert_duration_formula(r))
    ws.cell(row=r, column=COL_SIGMA, value=_sigma_formula(r))
    ws.cell(row=r, column=COL_BILL_E, value=_sum_of_rows("S", all_leaf_rows))


def _write_total_row(ws: Worksheet, r: int, phase_rows: list[int]) -> None:
    ws.cell(row=r, column=COL_ID, value="TOTAL")

    for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "S"):
        ws.cell(row=r, column=_col_idx(col),
                value=_sum_of_rows(col, phase_rows))


def _col_idx(letter: str) -> int:
    """Convert a single column letter (A-Z) to a 1-based index."""
    return ord(letter.upper()) - ord("A") + 1
