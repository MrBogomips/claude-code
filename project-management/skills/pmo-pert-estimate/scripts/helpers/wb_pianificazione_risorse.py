"""Resource Plan ("Pianificazione Risorse") sheet generator.

Produces a role × week PD matrix that supersedes the legacy Resources and
Timeline sheets. Each cell carries the person-day load contributed by a
primary role to a given calendar week. Aggregation rules:

* Primary role of a leaf activity = ``activity.resources[0]``. Activities with
  an empty resources list are skipped (and reported in the returned info dict).
* PERT effort of a leaf = ``(best + 4*likely + worst) / 6`` (PD).
* Per phase, per primary role: total PD = Σ(PERT) of activities mapping to
  that role. Distributed uniformly across the phase's calendar weeks.
* Phase weeks come from ``phase.start_week`` / ``phase.end_week`` when given,
  otherwise computed sequentially using a duration heuristic.

The output sheet is structured as follows::

    Row 1:           Header (Role | Code | Type | W1 | W2 | ... | TOTAL (PD))
    Row 2:           Calendar dates (W1 → project_start_date, +7d per col)
    Row 3..R:        One row per active role
    Row R+1:         Weekly TOTAL row (=SUM per week column)
    Row R+3+:        Capacity Warnings section (optional)

Cells are populated with raw numbers (PD), never percentages. The TOTAL (PD)
column carries a =SUM(...) formula so user edits propagate. Overcommitted
cells (>= 5 PD/week for a single role) get a red fill.
"""
from __future__ import annotations

import math
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from helpers.formatting import (
    HEADER_FONT,
    NUMBER_FMT,
    apply_column_widths,
    apply_style,
    get_total_style,
)
from helpers.i18n import t

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

FIXED_COL_COUNT = 3  # A=Role, B=Code, C=Type
OVERCOMMIT_FILL = PatternFill(fill_type="solid", fgColor="FFFFC7CE")  # light red
WARN_FILL = PatternFill(fill_type="solid", fgColor="FFFFEB9C")  # light yellow

#: PD/week considered the saturation threshold for a single role.
ROLE_WEEKLY_CAPACITY_PD = 5.0
#: Fraction of the saturation threshold that triggers the "warning" (yellow) fill.
WARN_FRACTION = 0.9


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build(wb: Workbook, data: dict, wbs_info: dict | None = None) -> dict:
    """Build the Resource Plan sheet and insert it into *wb*.

    Args:
        wb:       An :class:`openpyxl.Workbook` (must not already contain the
                  sheet under the localized name).
        data:     Input data dict (`config`, `phases`, `roles`).
        wbs_info: Optional WBS sheet info (unused today; reserved for future
                  cross-sheet references).

    Returns:
        A sheet_info dict::

            {
                "sheet_name":         "Pianificazione Risorse",
                "role_codes":         [...],
                "row_by_role":        {code: row_number},
                "week_col_start":     int,
                "total_col":          int,
                "calendar_row":       int,
                "weekly_total_row":   int,
                "total_weeks":        int,
                "role_total_pd":      {code: float},
                "skipped_activities": [activity_id, ...],
                "overcommits":        [(role_code, week_idx, pd, capacity), ...],
            }
    """
    config = data.get("config", {})
    phases = data.get("phases", [])
    roles_meta = {r["code"]: r for r in data.get("roles", [])}
    lang = config.get("lang", "en")

    # --- 1. Build the calendar plan: phase_id -> [week_indices] ---
    plan, total_weeks = _build_calendar_plan(phases, config)

    # --- 2. Compute PD allocations per (phase, primary_role) ---
    phase_role_pd, skipped = _compute_phase_role_pd(phases)

    # --- 3. Roll up into role × week matrix ---
    role_week_pd: dict[str, dict[int, float]] = {}
    for (phase_id, role_code), pd in phase_role_pd.items():
        weeks = plan.get(phase_id, [])
        if not weeks:
            continue
        share = pd / len(weeks)
        for w in weeks:
            role_week_pd.setdefault(role_code, {}).setdefault(w, 0.0)
            role_week_pd[role_code][w] += share

    role_codes = sorted(role_week_pd.keys())

    # --- 4. Create sheet and write contents ---
    sheet_name = t(lang, "sheet_resource_plan")
    ws: Worksheet = wb.create_sheet(sheet_name)

    week_col_start = FIXED_COL_COUNT + 1  # col D
    total_col = week_col_start + total_weeks  # one column after last week

    _write_header(ws, lang, total_weeks, week_col_start, total_col)
    calendar_row = 2
    _write_calendar_row(ws, calendar_row, config, total_weeks, week_col_start, lang)

    row_by_role: dict[str, int] = {}
    overcommits: list[tuple[str, int, float, float]] = []
    role_total_pd: dict[str, float] = {}

    data_start_row = 3
    current_row = data_start_row
    for code in role_codes:
        row_by_role[code] = current_row
        meta = roles_meta.get(code, {})
        role_label = meta.get("name", code)
        kind = t(lang, "rp_billable") if meta.get("billable", True) else t(lang, "rp_non_billable")
        ws.cell(row=current_row, column=1, value=role_label)
        ws.cell(row=current_row, column=2, value=code)
        ws.cell(row=current_row, column=3, value=kind)

        for w in range(1, total_weeks + 1):
            pd_value = role_week_pd[code].get(w, 0.0)
            col = week_col_start + (w - 1)
            cell = ws.cell(row=current_row, column=col, value=round(pd_value, 2) if pd_value else 0.0)
            cell.number_format = NUMBER_FMT
            if pd_value > ROLE_WEEKLY_CAPACITY_PD:
                cell.fill = OVERCOMMIT_FILL
                overcommits.append((code, w, pd_value, ROLE_WEEKLY_CAPACITY_PD))
            elif pd_value >= ROLE_WEEKLY_CAPACITY_PD * WARN_FRACTION:
                cell.fill = WARN_FILL

        # TOTAL (PD) column — SUM formula across the week columns
        total_cell = ws.cell(row=current_row, column=total_col,
                             value=_sum_range_formula(week_col_start, total_col - 1, current_row))
        total_cell.number_format = NUMBER_FMT
        total_cell.font = Font(bold=True)
        role_total_pd[code] = sum(role_week_pd[code].values())
        current_row += 1

    weekly_total_row = current_row
    _write_weekly_total_row(ws, weekly_total_row, lang, data_start_row,
                            current_row - 1, week_col_start, total_weeks, total_col)
    current_row += 1

    # Capacity Warnings block
    if overcommits:
        current_row += 1  # blank row
        warn_row = current_row
        ws.cell(row=warn_row, column=1, value=t(lang, "rp_capacity_warning")).font = Font(bold=True, color="FF9C0006")
        current_row += 1
        for code, w, pd, cap in overcommits:
            wk_label = f"W{w} ({code})"
            ws.cell(row=current_row, column=1,
                    value=t(lang, "rp_overcommit", wk=wk_label, pd=pd, cap=cap))
            current_row += 1

    # Column widths
    widths = {
        get_column_letter(1): 26,
        get_column_letter(2): 8,
        get_column_letter(3): 14,
    }
    for w in range(total_weeks):
        widths[get_column_letter(week_col_start + w)] = 10
    widths[get_column_letter(total_col)] = 14
    apply_column_widths(ws, widths)

    return {
        "sheet_name": sheet_name,
        "role_codes": role_codes,
        "row_by_role": row_by_role,
        "week_col_start": week_col_start,
        "total_col": total_col,
        "calendar_row": calendar_row,
        "weekly_total_row": weekly_total_row,
        "total_weeks": total_weeks,
        "role_total_pd": role_total_pd,
        "skipped_activities": skipped,
        "overcommits": overcommits,
    }


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------

def _pert_pd(activity: dict) -> float:
    return (activity["best_effort"] + 4 * activity["likely_effort"] + activity["worst_effort"]) / 6.0


def _phase_pert_duration_days(phase: dict) -> float:
    """PERT duration for a phase = SUM of leaf PERT durations (PD-equivalent)."""
    total = 0.0
    for wp in phase.get("work_packages", []):
        for a in wp.get("activities", []):
            total += (a["best_duration"] + 4 * a["likely_duration"] + a["worst_duration"]) / 6.0
    return total


def _phase_weeks(phase: dict, cursor: int) -> tuple[list[int], int]:
    """Return the week indices covered by *phase* and the next cursor.

    If start_week/end_week are present, use them. Otherwise, allocate sequentially
    starting at *cursor* with a length derived from PERT duration (ceil(d/5)).
    """
    if "start_week" in phase and "end_week" in phase:
        s, e = int(phase["start_week"]), int(phase["end_week"])
        return list(range(s, e + 1)), max(cursor, e + 1)
    duration_d = _phase_pert_duration_days(phase)
    weeks_needed = max(1, math.ceil(duration_d / 5.0))
    start = cursor
    end = start + weeks_needed - 1
    return list(range(start, end + 1)), end + 1


def _build_calendar_plan(phases: list[dict], config: dict) -> tuple[dict[str, list[int]], int]:
    plan: dict[str, list[int]] = {}
    cursor = 1
    max_week = 0
    for phase in phases:
        weeks, cursor = _phase_weeks(phase, cursor)
        plan[phase["id"]] = weeks
        if weeks:
            max_week = max(max_week, weeks[-1])
    total_weeks = config.get("calendar_total_weeks") or max_week or 1
    total_weeks = max(int(total_weeks), max_week)
    return plan, total_weeks


def _compute_phase_role_pd(phases: list[dict]) -> tuple[dict[tuple[str, str], float], list[str]]:
    out: dict[tuple[str, str], float] = {}
    skipped: list[str] = []
    for phase in phases:
        for wp in phase.get("work_packages", []):
            for a in wp.get("activities", []):
                resources = a.get("resources") or []
                if not resources:
                    skipped.append(a["id"])
                    continue
                primary = resources[0]
                pd = _pert_pd(a)
                key = (phase["id"], primary)
                out[key] = out.get(key, 0.0) + pd
    return out, skipped


def _write_header(ws: Worksheet, lang: str, total_weeks: int,
                  week_col_start: int, total_col: int) -> None:
    ws.cell(row=1, column=1, value=t(lang, "rp_role")).font = HEADER_FONT
    ws.cell(row=1, column=2, value=t(lang, "rp_code")).font = HEADER_FONT
    ws.cell(row=1, column=3, value=t(lang, "rp_type")).font = HEADER_FONT
    for w in range(1, total_weeks + 1):
        c = ws.cell(row=1, column=week_col_start + (w - 1), value=f"W{w}")
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    cell = ws.cell(row=1, column=total_col, value=t(lang, "rp_total_pd"))
    cell.font = HEADER_FONT


def _write_calendar_row(ws: Worksheet, row: int, config: dict, total_weeks: int,
                        week_col_start: int, lang: str) -> None:
    start_date_raw = config.get("project_start_date") or config.get("start_date")
    if not start_date_raw:
        return
    if isinstance(start_date_raw, str):
        try:
            d = datetime.fromisoformat(start_date_raw).date()
        except ValueError:
            return
    elif isinstance(start_date_raw, date):
        d = start_date_raw
    else:
        return
    label_cell = ws.cell(row=row, column=3, value=t(lang, "rp_calendar_row"))
    label_cell.font = Font(italic=True)
    for w in range(total_weeks):
        ws.cell(row=row, column=week_col_start + w, value=d.isoformat()).font = Font(italic=True)
        d += timedelta(days=7)


def _sum_range_formula(start_col: int, end_col: int, row: int) -> str:
    return f"=SUM({get_column_letter(start_col)}{row}:{get_column_letter(end_col)}{row})"


def _write_weekly_total_row(ws: Worksheet, row: int, lang: str,
                             data_start: int, data_end: int,
                             week_col_start: int, total_weeks: int,
                             total_col: int) -> None:
    ws.cell(row=row, column=1, value=t(lang, "rp_week_total"))
    total_style = get_total_style()
    apply_style(ws.cell(row=row, column=1), total_style)
    for w in range(total_weeks):
        col = week_col_start + w
        if data_end >= data_start:
            formula = f"=SUM({get_column_letter(col)}{data_start}:{get_column_letter(col)}{data_end})"
        else:
            formula = "=0"
        cell = ws.cell(row=row, column=col, value=formula)
        cell.number_format = NUMBER_FMT
        apply_style(cell, total_style)
    grand_cell = ws.cell(row=row, column=total_col,
                         value=_sum_range_formula(week_col_start, total_col - 1, row))
    grand_cell.number_format = NUMBER_FMT
    apply_style(grand_cell, total_style)
