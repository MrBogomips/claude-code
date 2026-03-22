"""Resources sheet generator for pmo-pert-estimate.

Builds the 'Resources' worksheet with:
  - Row 1 (metadata): billable flags ('Y'/'N') in role columns
  - Row 2 (header):   Phase | Description | Team | [role names] | TOTAL EFFORT | BILLABLE EFFORT
  - Data rows:        one row per unique (phase, team) allocation entry
  - Team subtotals:   one subtotal row per distinct team
  - Footer row:       grand total SUM formulas

All rows in the data range share the same TOTAL EFFORT and BILLABLE EFFORT
formula patterns, enabling consistent cross-column summation.
"""
from __future__ import annotations

from collections import defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from helpers.formatting import (
    apply_style,
    get_formula_fill,
    get_total_style,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_FIXED_COLS = 3          # A=Phase, B=Description, C=Team
_DATA_START_ROW = 3      # first data row (after metadata + header)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build(wb: Workbook, data: dict) -> dict:
    """Add the Resources sheet to *wb* and return sheet_info.

    Args:
        wb:   An :class:`openpyxl.Workbook` instance (must not already contain
              a sheet named ``"Resources"``).
        data: Input dict with keys:
              - ``"config"``: project configuration dict
              - ``"roles"``: list of role dicts (``code``, ``name``, ``billable``)
              - ``"resource_allocation"``: list of allocation dicts
              - ``"phases"``: list of phase dicts (not used directly here)

    Returns:
        A ``sheet_info`` dict with:
        - ``"sheet_name"``: ``"Resources"``
        - ``"role_columns"``: mapping of role_code → column letter
        - ``"first_role_col"``: first role column letter
        - ``"last_role_col"``: last role column letter
        - ``"total_effort_col"``: column letter for TOTAL EFFORT
        - ``"billable_effort_col"``: column letter for BILLABLE EFFORT
        - ``"data_start_row"``: 3
        - ``"data_end_row"``: last data/subtotal row number
        - ``"team_subtotal_rows"``: mapping of team_name → row number
    """
    roles: list[dict] = data["roles"]
    allocations: list[dict] = data["resource_allocation"]

    ws = wb.create_sheet("Resources")

    # ------------------------------------------------------------------
    # Compute column layout
    # ------------------------------------------------------------------
    # Fixed: A=Phase (1), B=Description (2), C=Team (3)
    # Roles: columns 4..(3+len(roles))
    # TOTAL EFFORT: column 4+len(roles)
    # BILLABLE EFFORT: column 5+len(roles)

    role_col_indices: dict[str, int] = {
        role["code"]: _FIXED_COLS + 1 + i
        for i, role in enumerate(roles)
    }
    first_role_idx = _FIXED_COLS + 1
    last_role_idx = _FIXED_COLS + len(roles)
    total_effort_idx = last_role_idx + 1
    billable_effort_idx = last_role_idx + 2

    first_role_col = get_column_letter(first_role_idx)
    last_role_col = get_column_letter(last_role_idx)
    total_effort_col = get_column_letter(total_effort_idx)
    billable_effort_col = get_column_letter(billable_effort_idx)

    role_columns: dict[str, str] = {
        code: get_column_letter(idx)
        for code, idx in role_col_indices.items()
    }

    # ------------------------------------------------------------------
    # Row 1: Metadata (billable flags)
    # ------------------------------------------------------------------
    _write_metadata_row(ws, roles, role_col_indices)

    # ------------------------------------------------------------------
    # Row 2: Header
    # ------------------------------------------------------------------
    _write_header_row(
        ws, roles, role_col_indices,
        total_effort_col, billable_effort_col,
    )

    # ------------------------------------------------------------------
    # Group allocations by team (preserving team insertion order)
    # ------------------------------------------------------------------
    teams_order: list[str] = []
    by_team: dict[str, list[dict]] = defaultdict(list)
    for alloc in allocations:
        team = alloc["team"]
        if team not in by_team:
            teams_order.append(team)
        by_team[team].append(alloc)

    # ------------------------------------------------------------------
    # Data rows + team subtotals
    # ------------------------------------------------------------------
    current_row = _DATA_START_ROW
    team_subtotal_rows: dict[str, int] = {}

    for team in teams_order:
        team_allocs = by_team[team]
        team_data_rows: list[int] = []

        for alloc in team_allocs:
            _write_data_row(
                ws=ws,
                row=current_row,
                alloc=alloc,
                role_col_indices=role_col_indices,
                first_role_col=first_role_col,
                last_role_col=last_role_col,
                total_effort_col=total_effort_col,
                billable_effort_col=billable_effort_col,
            )
            team_data_rows.append(current_row)
            current_row += 1

        # Team subtotal row
        subtotal_row = current_row
        team_subtotal_rows[team] = subtotal_row
        _write_subtotal_row(
            ws=ws,
            row=subtotal_row,
            team=team,
            team_data_rows=team_data_rows,
            role_col_indices=role_col_indices,
            first_role_col=first_role_col,
            last_role_col=last_role_col,
            total_effort_col=total_effort_col,
            billable_effort_col=billable_effort_col,
        )
        current_row += 1

    data_end_row = current_row - 1

    # ------------------------------------------------------------------
    # Footer: grand total (sums team subtotals to avoid double-counting)
    # ------------------------------------------------------------------
    footer_row = data_end_row + 1
    _write_footer_row(
        ws=ws,
        row=footer_row,
        team_subtotal_rows=team_subtotal_rows,
        total_effort_col=total_effort_col,
        billable_effort_col=billable_effort_col,
    )

    return {
        "sheet_name": "Resources",
        "role_columns": role_columns,
        "first_role_col": first_role_col,
        "last_role_col": last_role_col,
        "total_effort_col": total_effort_col,
        "billable_effort_col": billable_effort_col,
        "data_start_row": _DATA_START_ROW,
        "data_end_row": data_end_row,
        "team_subtotal_rows": team_subtotal_rows,
    }


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _write_metadata_row(
    ws,
    roles: list[dict],
    role_col_indices: dict[str, int],
) -> None:
    """Write Row 1: billable flags ('Y'/'N') in role columns."""
    italic_font = Font(italic=True, size=9)
    fill = PatternFill(fill_type="solid", fgColor="FFFFF2CC")

    for role in roles:
        col_idx = role_col_indices[role["code"]]
        cell = ws.cell(row=1, column=col_idx)
        cell.value = "Y" if role["billable"] else "N"
        cell.font = italic_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _write_header_row(
    ws,
    roles: list[dict],
    role_col_indices: dict[str, int],
    total_effort_col: str,
    billable_effort_col: str,
) -> None:
    """Write Row 2: column headers."""
    bold_font = Font(bold=True, size=11)
    fill = PatternFill(fill_type="solid", fgColor="FFD9E2F3")
    wrap = Alignment(wrap_text=True, horizontal="center")

    fixed_headers = {1: "Phase", 2: "Description", 3: "Team"}
    for col_idx, label in fixed_headers.items():
        cell = ws.cell(row=2, column=col_idx)
        cell.value = label
        cell.font = bold_font
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True)

    for role in roles:
        col_idx = role_col_indices[role["code"]]
        cell = ws.cell(row=2, column=col_idx)
        cell.value = role["name"]
        cell.font = bold_font
        cell.fill = fill
        cell.alignment = wrap

    for col_letter, label in (
        (total_effort_col, "TOTAL EFFORT"),
        (billable_effort_col, "BILLABLE EFFORT"),
    ):
        cell = ws[f"{col_letter}2"]
        cell.value = label
        cell.font = bold_font
        cell.fill = fill
        cell.alignment = wrap


def _total_effort_formula(first_role: str, last_role: str, row: int) -> str:
    return f"=SUM({first_role}{row}:{last_role}{row})"


def _billable_effort_formula(first_role: str, last_role: str, row: int) -> str:
    return (
        f"=SUMPRODUCT({first_role}{row}:{last_role}{row},"
        f'({first_role}$1:{last_role}$1="Y")*1)'
    )


def _write_data_row(
    ws,
    row: int,
    alloc: dict,
    role_col_indices: dict[str, int],
    first_role_col: str,
    last_role_col: str,
    total_effort_col: str,
    billable_effort_col: str,
) -> None:
    """Write one leaf data row for a single allocation entry."""
    normal_font = Font(bold=False)
    wrap = Alignment(wrap_text=True)

    ws.cell(row=row, column=1).value = alloc["phase_name"]
    ws.cell(row=row, column=2).value = alloc.get("description", "")
    ws.cell(row=row, column=3).value = alloc["team"]

    for col in (1, 2, 3):
        cell = ws.cell(row=row, column=col)
        cell.font = normal_font
        cell.alignment = wrap

    # Effort value in the appropriate role column
    role_code = alloc["role_code"]
    if role_code in role_col_indices:
        col_idx = role_col_indices[role_code]
        cell = ws.cell(row=row, column=col_idx)
        cell.value = alloc["effort"]
        cell.font = normal_font

    # Formulas
    formula_fill = get_formula_fill()
    for col_letter, formula in (
        (total_effort_col, _total_effort_formula(first_role_col, last_role_col, row)),
        (billable_effort_col, _billable_effort_formula(first_role_col, last_role_col, row)),
    ):
        cell = ws[f"{col_letter}{row}"]
        cell.value = formula
        cell.font = normal_font
        cell.fill = formula_fill


def _write_subtotal_row(
    ws,
    row: int,
    team: str,
    team_data_rows: list[int],
    role_col_indices: dict[str, int],
    first_role_col: str,
    last_role_col: str,
    total_effort_col: str,
    billable_effort_col: str,
) -> None:
    """Write a team subtotal row with SUM formulas for each role column."""
    subtotal_font = Font(bold=True)
    fill = PatternFill(fill_type="solid", fgColor="FFD9E2F3")
    wrap = Alignment(wrap_text=True)

    ws.cell(row=row, column=1).value = f"Subtotal — {team}"
    ws.cell(row=row, column=2).value = ""
    ws.cell(row=row, column=3).value = team

    for col in (1, 2, 3):
        cell = ws.cell(row=row, column=col)
        cell.font = subtotal_font
        cell.fill = fill
        cell.alignment = wrap

    # SUM formula per role column
    for col_idx in role_col_indices.values():
        col_letter = get_column_letter(col_idx)
        refs = ",".join(f"{col_letter}{r}" for r in team_data_rows)
        cell = ws.cell(row=row, column=col_idx)
        cell.value = f"=SUM({refs})"
        cell.font = subtotal_font
        cell.fill = fill

    # TOTAL EFFORT and BILLABLE EFFORT (same formula pattern as data rows)
    formula_fill = get_formula_fill()
    for col_letter, formula in (
        (total_effort_col, _total_effort_formula(first_role_col, last_role_col, row)),
        (billable_effort_col, _billable_effort_formula(first_role_col, last_role_col, row)),
    ):
        cell = ws[f"{col_letter}{row}"]
        cell.value = formula
        cell.font = subtotal_font
        cell.fill = formula_fill


def _write_footer_row(
    ws,
    row: int,
    team_subtotal_rows: dict[str, int],
    total_effort_col: str,
    billable_effort_col: str,
) -> None:
    """Write the grand total footer row.

    Sums only team subtotal rows to avoid double-counting data rows
    that are already included in those subtotals.
    """
    style = get_total_style()

    ws.cell(row=row, column=1).value = "TOTAL"
    ws.cell(row=row, column=2).value = ""
    ws.cell(row=row, column=3).value = ""

    for col in (1, 2, 3):
        apply_style(ws.cell(row=row, column=col), style)

    formula_fill = get_formula_fill()

    subtotal_row_nums = sorted(team_subtotal_rows.values())
    refs_total = ",".join(f"{total_effort_col}{r}" for r in subtotal_row_nums)
    refs_billable = ",".join(f"{billable_effort_col}{r}" for r in subtotal_row_nums)

    total_formula = f"=SUM({refs_total})"
    cell = ws[f"{total_effort_col}{row}"]
    cell.value = total_formula
    apply_style(cell, style)
    cell.fill = formula_fill

    billable_formula = f"=SUM({refs_billable})"
    cell = ws[f"{billable_effort_col}{row}"]
    cell.value = billable_formula
    apply_style(cell, style)
    cell.fill = formula_fill
