"""Formatting helpers for pmo-pert-estimate Excel output.

Each get_*_style() function creates and returns a NEW dict containing
openpyxl style objects (Font, PatternFill, Border, Alignment).  Callers
apply the dict to a cell via apply_style().  No NamedStyle objects are
used so there is no risk of double-registration with a workbook.
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

#: Reusable header font — bold, size 11.  Read-only; never mutate.
HEADER_FONT: Font = Font(bold=True, size=11)

#: Standard number format for effort / duration values.
NUMBER_FMT: str = "#,##0.00"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def get_number_format() -> str:
    """Return the standard number format string."""
    return NUMBER_FMT


def get_phase_style(primary_color: str) -> dict:
    """Return a style dict for a Phase header row.

    Args:
        primary_color: Hex colour string without leading '#', e.g. ``"1B4FA5"``.

    Returns:
        A fresh dict with keys ``font``, ``fill``, and ``alignment``.
    """
    return {
        "font": Font(bold=True, color="FFFFFFFF"),
        "fill": PatternFill(
            fill_type="solid",
            fgColor=f"FF{primary_color.upper()}",
        ),
        "alignment": Alignment(wrap_text=True),
    }


def get_wp_style() -> dict:
    """Return a style dict for a Work-Package header row.

    Returns:
        A fresh dict with keys ``font``, ``fill``, and ``alignment``.
    """
    return {
        "font": Font(bold=True),
        "fill": PatternFill(fill_type="solid", fgColor="FFD9E2F3"),
        "alignment": Alignment(wrap_text=True),
    }


def get_leaf_style() -> dict:
    """Return a style dict for a leaf Activity row (normal, no background fill).

    Returns:
        A fresh dict with keys ``font`` and ``alignment``.
    """
    return {
        "font": Font(bold=False),
        "alignment": Alignment(wrap_text=True),
    }


def get_total_style() -> dict:
    """Return a style dict for a TOTAL / summary row.

    Characteristics:
    - Bold white text on dark (``2F2F2F``) background.
    - Double top border.

    Returns:
        A fresh dict with keys ``font``, ``fill``, ``border``, and ``alignment``.
    """
    return {
        "font": Font(bold=True, color="FFFFFFFF"),
        "fill": PatternFill(fill_type="solid", fgColor="FF2F2F2F"),
        "border": Border(top=Side(border_style="double")),
        "alignment": Alignment(wrap_text=True),
    }


def get_formula_fill() -> PatternFill:
    """Return a PatternFill for formula / computed-value columns (``FFF2CC``).

    Returns:
        A fresh :class:`openpyxl.styles.PatternFill` instance.
    """
    return PatternFill(fill_type="solid", fgColor="FFFFF2CC")


def apply_column_widths(ws, widths: dict) -> None:
    """Set column widths on a worksheet.

    Args:
        ws:     An :class:`openpyxl.worksheet.worksheet.Worksheet` instance.
        widths: Mapping of column letter to desired width, e.g.
                ``{"A": 20, "B": 35}``.  The input dict is never mutated.
    """
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def apply_style(cell, style_dict: dict) -> None:
    """Apply a style dictionary to a cell.

    Args:
        cell:       An openpyxl cell object.
        style_dict: A dict whose keys are cell attribute names
                    (``"font"``, ``"fill"``, ``"border"``, ``"alignment"``)
                    and whose values are the corresponding openpyxl objects.
    """
    for attr, value in style_dict.items():
        setattr(cell, attr, value)
