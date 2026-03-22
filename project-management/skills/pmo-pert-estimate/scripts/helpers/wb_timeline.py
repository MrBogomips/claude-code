"""Timeline/Gantt sheet generator for pmo-pert-estimate Excel output.

Builds a "Timeline" sheet that cross-references the WBS sheet and renders
a Gantt chart with critical-path colouring.

Algorithm overview
------------------
1. Extract all leaf activities from ``data["phases"]``, compute PERT duration
   for each: ``(best + 4*likely + worst) / 6``.
2. **Forward pass** (PERT CPM):
   ES = max(EF of predecessors), EF = ES + pert_duration.
   Activities with no predecessors start at ES = 0.
3. **Backward pass**:
   LF = min(LS of successors), LS = LF - pert_duration.
   Total Float = LS - ES.  Float == 0  →  on critical path.
4. **Activity classification**:
   - float == 0                               → critical (red, ``FF0000``)
   - has dependencies AND float > 0           → parallel  (blue, ``4472C4``)
   - no deps AND no dependents                → continuous (orange, ``FFA500``)
5. **Period columns**: period_length = 10 (biweekly) / 5 (weekly) / 20 (monthly).
   num_periods = ceil(total_duration / period_length).
6. **Sheet layout**:
   Row 1 : header (ID, Phase/WP/Activity, PERT Duration, σ, P1, P2, …)
   Rows 2…N: one row per WBS data row (phases, WPs, leaves).
             Cols A-D: cross-refs to WBS sheet.
             Period columns for leaf rows only: coloured blocks.
   Summary block, then legend.
"""
from __future__ import annotations

import math
from datetime import date, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from helpers.formatting import HEADER_FONT, apply_column_widths, apply_style


# ---------------------------------------------------------------------------
# Colour constants (openpyxl uses ARGB with leading FF)
# ---------------------------------------------------------------------------

COLOR_CRITICAL = "FFFF0000"    # Red   – critical path
COLOR_PARALLEL = "FF4472C4"    # Blue  – parallel (has deps, not critical)
COLOR_CONTINUOUS = "FFFFA500"  # Orange – no deps, no dependents


def _make_fill(argb: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def build(wb: Workbook, data: dict, wbs_info: dict) -> dict:
    """Build the Timeline/Gantt sheet and insert it into *wb*.

    Args:
        wb:       An :class:`openpyxl.Workbook` instance.
        data:     Input dict with keys ``config``, ``phases``, ``roles``.
        wbs_info: The dict returned by :func:`helpers.wb_wbs.build`.

    Returns:
        A ``sheet_info`` dict::

            {
                "sheet_name":       "Timeline",
                "num_periods":      N,
                "summary_start_row": R,
            }
    """
    config = data["config"]
    phases = data["phases"]

    period_type = config.get("period_type", "biweekly")
    period_length = _period_length(period_type)
    start_date_str: Optional[str] = config.get("start_date")

    # --- Build schedule (CPM) ---
    activities = _collect_activities(phases)
    schedule = _compute_schedule(activities)
    critical_ids = _critical_path_ids(schedule)
    continuous_ids = _continuous_ids(schedule)

    total_duration = max(
        (info["ef"] for info in schedule.values()),
        default=0.0,
    )
    num_periods = max(1, math.ceil(total_duration / period_length))

    ws: Worksheet = wb.create_sheet("Timeline")

    # --- Header ---
    _write_header(ws, num_periods, period_type)

    # --- Data rows (mirror WBS layout) ---
    data_start = wbs_info["data_start_row"]
    data_end = wbs_info["data_end_row"]

    for r in range(data_start, data_end + 1):
        _write_data_row(ws, r, num_periods)

    # Fill Gantt cells for leaf activities
    _fill_gantt_cells(
        ws, schedule, critical_ids, continuous_ids,
        data_start, data_end, num_periods, period_length,
        activities_by_id={a["id"]: a for a in activities},
        wbs_sheet=wb["WBS"],
    )

    # --- Summary block ---
    summary_start = data_end + 2  # leave one blank row
    summary_end = _write_summary(
        ws, summary_start, schedule, wbs_info, num_periods,
        period_length, start_date_str,
    )

    # --- Legend ---
    legend_start = summary_end + 2
    _write_legend(ws, legend_start)

    # --- Column widths ---
    col_widths = {"A": 8, "B": 40, "C": 14, "D": 12}
    for p in range(1, num_periods + 1):
        col_widths[get_column_letter(4 + p)] = 6
    apply_column_widths(ws, col_widths)

    return {
        "sheet_name": "Timeline",
        "num_periods": num_periods,
        "summary_start_row": summary_start,
    }


# ---------------------------------------------------------------------------
# Schedule computation helpers
# ---------------------------------------------------------------------------

def _collect_activities(phases: list[dict]) -> list[dict]:
    """Flatten all leaf activities from phases/work_packages."""
    result = []
    for phase in phases:
        for wp in phase.get("work_packages", []):
            for act in wp.get("activities", []):
                pert_d = _pert(
                    act["best_duration"],
                    act["likely_duration"],
                    act["worst_duration"],
                )
                result.append({
                    "id": act["id"],
                    "name": act["name"],
                    "pert_duration": pert_d,
                    "sigma": (act["worst_duration"] - act["best_duration"]) / 6,
                    "dependencies": list(act.get("dependencies", [])),
                })
    return result


def _pert(best: float, likely: float, worst: float) -> float:
    return (best + 4 * likely + worst) / 6


def _compute_schedule(activities: list[dict]) -> dict:
    """Compute ES, EF, LS, LF, float for all activities.

    Continuous activities (no dependencies AND not a prerequisite for anything)
    are identified first and excluded from the CPM network.  Their ES is set to
    0 and EF to the project end determined by the remaining network.  This
    ensures the critical path reflects the actual dependency chain, not the
    longest standalone activity.

    Returns a dict keyed by activity id::

        {
            id: {
                "es": float, "ef": float,
                "ls": float, "lf": float,
                "float": float,
                "pert_duration": float,
                "sigma": float,
                "dependencies": list[str],
                "is_continuous": bool,
            }
        }
    """
    acts_by_id = {a["id"]: a for a in activities}

    # Identify which activity ids are depended upon by others
    all_dep_targets: set[str] = set()
    for a in activities:
        all_dep_targets.update(a["dependencies"])

    # Continuous: no deps AND nothing depends on them
    continuous_activity_ids: set[str] = {
        a["id"] for a in activities
        if not a["dependencies"] and a["id"] not in all_dep_targets
    }

    # Network activities (exclude continuous from CPM)
    network_acts = [a for a in activities if a["id"] not in continuous_activity_ids]
    network_ids = {a["id"] for a in network_acts}

    # Build adjacency for network activities only
    in_degree: dict[str, int] = {a["id"]: 0 for a in network_acts}
    successors: dict[str, list[str]] = {a["id"]: [] for a in network_acts}
    predecessors: dict[str, list[str]] = {a["id"]: [] for a in network_acts}

    for a in network_acts:
        for dep in a["dependencies"]:
            if dep in in_degree:
                in_degree[a["id"]] += 1
                successors[dep].append(a["id"])
                predecessors[a["id"]].append(dep)

    # Topological sort (Kahn's algorithm)
    queue = [nid for nid in in_degree if in_degree[nid] == 0]
    topo_order: list[str] = []
    remaining = dict(in_degree)
    while queue:
        queue.sort()
        nid = queue.pop(0)
        topo_order.append(nid)
        for succ in successors[nid]:
            remaining[succ] -= 1
            if remaining[succ] == 0:
                queue.append(succ)

    # Forward pass
    es: dict[str, float] = {}
    ef: dict[str, float] = {}
    for nid in topo_order:
        act = acts_by_id[nid]
        if not predecessors[nid]:
            es[nid] = 0.0
        else:
            es[nid] = max(ef[p] for p in predecessors[nid] if p in ef)
        ef[nid] = es[nid] + act["pert_duration"]

    # Project end = max EF in the network (not counting continuous activities)
    network_project_end = max(ef.values()) if ef else 0.0

    # Backward pass
    lf: dict[str, float] = {}
    ls: dict[str, float] = {}
    for nid in reversed(topo_order):
        act = acts_by_id[nid]
        if not successors[nid]:
            lf[nid] = network_project_end
        else:
            lf[nid] = min(ls[s] for s in successors[nid] if s in ls)
        ls[nid] = lf[nid] - act["pert_duration"]

    # Build schedule dict — network activities
    schedule: dict[str, dict] = {}
    for a in network_acts:
        nid = a["id"]
        total_float = ls.get(nid, 0.0) - es.get(nid, 0.0)
        schedule[nid] = {
            "es": es.get(nid, 0.0),
            "ef": ef.get(nid, 0.0),
            "ls": ls.get(nid, 0.0),
            "lf": lf.get(nid, 0.0),
            "float": round(total_float, 10),
            "pert_duration": a["pert_duration"],
            "sigma": a["sigma"],
            "dependencies": a["dependencies"],
            "is_continuous": False,
        }

    # Continuous activities span full project (ES=0, EF=project_end)
    overall_end = max(
        network_project_end,
        max(
            (a["pert_duration"] for a in activities if a["id"] in continuous_activity_ids),
            default=0.0,
        ),
    )
    for a in activities:
        if a["id"] in continuous_activity_ids:
            schedule[a["id"]] = {
                "es": 0.0,
                "ef": overall_end,
                "ls": 0.0,
                "lf": overall_end,
                "float": 0.0,  # spans full project, treated specially
                "pert_duration": a["pert_duration"],
                "sigma": a["sigma"],
                "dependencies": a["dependencies"],
                "is_continuous": True,
            }

    return schedule


def _critical_path_ids(schedule: dict) -> set[str]:
    """Return set of activity ids on the critical path (total float ≈ 0, not continuous)."""
    return {
        nid for nid, info in schedule.items()
        if abs(info["float"]) < 1e-6 and not info.get("is_continuous", False)
    }


def _continuous_ids(schedule: dict) -> set[str]:
    """Return ids of activities flagged as continuous (no deps, no dependents)."""
    return {nid for nid, info in schedule.items() if info.get("is_continuous", False)}


# ---------------------------------------------------------------------------
# Sheet writing helpers
# ---------------------------------------------------------------------------

def _period_length(period_type: str) -> int:
    """Return number of working days per period."""
    return {"biweekly": 10, "weekly": 5, "monthly": 20}.get(period_type, 10)


def _period_prefix(period_type: str) -> str:
    """Return the header prefix for period columns based on period type."""
    return {"biweekly": "P", "weekly": "W", "monthly": "M"}.get(period_type, "P")


def _write_header(ws: Worksheet, num_periods: int, period_type: str = "biweekly") -> None:
    headers = ["ID", "Phase/WP/Activity", "PERT Duration", "\u03c3"]
    prefix = _period_prefix(period_type)
    for p in range(1, num_periods + 1):
        headers.append(f"{prefix}{p}")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT


def _write_data_row(ws: Worksheet, r: int, num_periods: int) -> None:
    """Write cross-reference formulas for cols A-D on a data row."""
    ws.cell(row=r, column=1, value=f"=WBS!A{r}")
    # Concatenate Phase (B), WP (C), Activity (D) — show whichever is non-empty
    concat = (
        f"=WBS!B{r}"
        f'&IF(WBS!C{r}<>"", " "&WBS!C{r}, "")'
        f'&IF(WBS!D{r}<>"", " "&WBS!D{r}, "")'
    )
    ws.cell(row=r, column=2, value=concat)
    ws.cell(row=r, column=3, value=f"=WBS!L{r}")
    ws.cell(row=r, column=4, value=f"=WBS!M{r}")


def _fill_gantt_cells(
    ws: Worksheet,
    schedule: dict,
    critical_ids: set[str],
    continuous_ids: set[str],
    data_start: int,
    data_end: int,
    num_periods: int,
    period_length: int,
    activities_by_id: dict,
    wbs_sheet: Worksheet,
) -> None:
    """Fill Gantt bar cells for each leaf activity."""
    period_col_start = 5  # column E

    for r in range(data_start, data_end + 1):
        act_id = wbs_sheet[f"A{r}"].value
        if act_id is None or act_id not in schedule:
            continue  # phase or WP row → skip

        info = schedule[act_id]
        es = info["es"]
        ef = info["ef"]

        # Determine colour
        if act_id in critical_ids:
            fill = _make_fill(COLOR_CRITICAL)
        elif act_id in continuous_ids:
            fill = _make_fill(COLOR_CONTINUOUS)
        else:
            fill = _make_fill(COLOR_PARALLEL)

        # Which periods does this activity span?
        for p in range(num_periods):
            period_start = p * period_length
            period_end = (p + 1) * period_length
            # Activity overlaps period if es < period_end AND ef > period_start
            if es < period_end and ef > period_start:
                col_letter = get_column_letter(period_col_start + p)
                ws[f"{col_letter}{r}"].fill = fill


def _write_summary(
    ws: Worksheet,
    summary_start: int,
    schedule: dict,
    wbs_info: dict,
    num_periods: int,
    period_length: int,
    start_date_str: Optional[str],
) -> int:
    """Write the summary block; return the last row used."""
    r = summary_start
    bold_font = Font(bold=True)

    # Critical Path row
    critical_ids = _critical_path_ids(schedule)
    cp_label = "Critical Path"
    cp_value = " → ".join(sorted(critical_ids)) if critical_ids else "N/A"
    _write_summary_row(ws, r, cp_label, cp_value, bold_font)
    r += 1

    # Total PERT Duration — reference WBS total row L column
    total_row = wbs_info["total_row"]
    _write_summary_row(ws, r, "Total PERT Duration", f"=WBS!L{total_row}", bold_font)
    r += 1

    # σ total (from WBS total row M column)
    sigma_total_formula = f"=WBS!M{total_row}"

    # CI 68%: μ ± σ  → show as [μ-σ, μ+σ]
    ci68_formula = f'=WBS!L{total_row}&" ± "&WBS!M{total_row}'
    _write_summary_row(ws, r, "CI 68%", ci68_formula, bold_font)
    r += 1

    # CI 95%: μ ± 2σ
    ci95_formula = f'=WBS!L{total_row}&" ± 2×"&WBS!M{total_row}'
    _write_summary_row(ws, r, "CI 95%", ci95_formula, bold_font)
    r += 1

    # Indicative dates (only when start_date configured)
    if start_date_str:
        start_dt = _parse_date(start_date_str)
        if start_dt is not None:
            _write_summary_row(ws, r, "Start Date", start_dt, bold_font)
            r += 1

            # End date: start + total_duration working days (approximate)
            total_pert = sum(
                info["pert_duration"] for info in schedule.values()
                if info["es"] == max(
                    (v["es"] for v in schedule.values()), default=0
                )
            ) if schedule else 0
            # Simpler: use num_periods * period_length as upper bound
            project_end_duration = max(
                (info["ef"] for info in schedule.values()), default=0.0
            )
            end_dt = _add_working_days(start_dt, math.ceil(project_end_duration))
            _write_summary_row(ws, r, "End Date", end_dt, bold_font)
            r += 1

    return r - 1


def _write_summary_row(
    ws: Worksheet, r: int, label: str, value, font: Font
) -> None:
    label_cell = ws.cell(row=r, column=1, value=label)
    label_cell.font = font
    value_cell = ws.cell(row=r, column=2, value=value)


def _write_legend(ws: Worksheet, start_row: int) -> None:
    """Write a 3-entry legend below the summary."""
    r = start_row
    bold = Font(bold=True)

    legend_header = ws.cell(row=r, column=1, value="Legend")
    legend_header.font = Font(bold=True, underline="single")
    r += 1

    entries = [
        ("Critical Path", COLOR_CRITICAL),
        ("Parallel Activity", COLOR_PARALLEL),
        ("Continuous Activity", COLOR_CONTINUOUS),
    ]
    for label, color in entries:
        color_cell = ws.cell(row=r, column=1)
        color_cell.fill = _make_fill(color)
        color_cell.value = ""
        label_cell = ws.cell(row=r, column=2, value=label)
        r += 1


# ---------------------------------------------------------------------------
# Date utilities
# ---------------------------------------------------------------------------

def _parse_date(date_str: str) -> Optional[date]:
    """Parse ISO date string, return None on failure."""
    try:
        parts = date_str.split("-")
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return None


def _add_working_days(start: date, days: int) -> date:
    """Add *days* working days (Mon-Fri) to *start*."""
    current = start
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:  # Mon=0 … Fri=4
            added += 1
    return current
