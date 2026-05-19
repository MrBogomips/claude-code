"""Summary sheet generator for pmo-pert-estimate.

Builds the Summary sheet of the 4-sheet workbook. Compared to the legacy
Summary, this version:

* Drops the sequential-sum CI Duration block (Issue #2): aggregating leaf
  durations sequentially ignores phase parallelism and produces misleading
  numbers. Calendar duration is now a single explicit value, derived from
  config.calendar_total_weeks or from phase.start_week/end_week.
* Adds PM/DevOps overhead rows applied on Tech PERT (Issue #5).
* Calculates Management Reserve on the correct base (Tech + Overhead +
  Contingency), realised as a formula referencing the Fascia BASSA cell
  (Issue #3).
* Surfaces three effort bands — Fascia BASSA, MEDIA (raccomandata), ALTA.
* Computes Effort by Team as literal PD aggregated from the WBS primary role
  of each leaf (Issue #1), not as a Resources!… cross-ref.
* Optionally lists sensitivity scenarios as plain text rows.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from helpers.formatting import (
    HEADER_FONT,
    NUMBER_FMT,
    apply_column_widths,
    apply_style,
    get_formula_fill,
    get_phase_style,
    get_total_style,
)
from helpers.i18n import t

# ---------------------------------------------------------------------------
# Phase table layout (unchanged)
# ---------------------------------------------------------------------------

_COL_PHASE = 1       # A
_COL_DESC = 2        # B
_COL_BEST_E = 3      # C
_COL_LIKELY_E = 4    # D
_COL_WORST_E = 5     # E
_COL_PERT_E = 6      # F
_COL_BEST_D = 7      # G
_COL_LIKELY_D = 8    # H
_COL_WORST_D = 9     # I
_COL_PERT_D = 10     # J
_COL_SIGMA = 11      # K

_HEADER_ROW = 1
_DATA_START_ROW = 2

_COLUMN_WIDTHS = {
    "A": 32,
    "B": 36,
    "C": 14,
    "D": 14,
    "E": 14,
    "F": 14,
    "G": 14,
    "H": 14,
    "I": 14,
    "J": 14,
    "K": 12,
}


def build(
    wb: Workbook,
    data: dict,
    wbs_info: dict,
    risks_info: dict,
    rp_info: dict,
) -> None:
    """Build the Summary sheet and add it to *wb*."""
    config = data["config"]
    phases = data["phases"]
    roles = data.get("roles", [])
    lang = config.get("lang", "en")
    primary_color = config.get("primary_color", "1B4FA5")
    effort_unit = config.get("effort_unit", "pd")
    duration_unit = config.get("duration_unit", "d")

    sheet_name = t(lang, "sheet_summary")
    ws: Worksheet = wb.create_sheet(sheet_name)

    # --- 1. Phase table ---
    _write_header(ws, effort_unit, duration_unit)
    phase_rows = wbs_info["phase_rows"]
    for offset, (phase, wbs_phase_row) in enumerate(zip(phases, phase_rows)):
        _write_phase_row(ws, _DATA_START_ROW + offset, phase, wbs_phase_row)

    data_start = _DATA_START_ROW
    data_end = _DATA_START_ROW + len(phases) - 1
    total_row = data_end + 1
    _write_total_row(ws, total_row, data_start, data_end)

    pert_e_col = get_column_letter(_COL_PERT_E)

    # --- 2. Effort breakdown block ---
    current = total_row + 2  # blank separator
    pm_pct = float(config.get("pm_overhead_pct") or 0)
    devops_pct = float(config.get("devops_overhead_pct") or 0)
    alta_pct = float(config.get("alta_uplift_pct") or 0)
    mr_pct = float(config.get("management_reserve_pct") or 0)

    # Tech PERT Effort
    tech_row = current
    ws.cell(row=tech_row, column=1, value=t(lang, "summary_tech_pert"))
    ws.cell(row=tech_row, column=2, value=f"={pert_e_col}{total_row}")
    current += 1

    # PM Overhead
    pm_row = current
    ws.cell(row=pm_row, column=1, value=t(lang, "summary_pm_overhead", pct=_pct_label(pm_pct)))
    ws.cell(row=pm_row, column=2, value=f"=B{tech_row}*{pm_pct}")
    current += 1

    # DevOps Overhead
    devops_row = current
    ws.cell(row=devops_row, column=1,
            value=t(lang, "summary_devops_overhead", pct=_pct_label(devops_pct)))
    ws.cell(row=devops_row, column=2, value=f"=B{tech_row}*{devops_pct}")
    current += 1

    # Subtotal Tech + Overhead
    subtotal_row = current
    ws.cell(row=subtotal_row, column=1, value=t(lang, "summary_subtotal"))
    ws.cell(row=subtotal_row, column=2, value=f"=B{tech_row}+B{pm_row}+B{devops_row}")
    current += 1

    # Contingency per-risk
    contingency_row = current
    risks_contingency_row = risks_info["total_contingency_row"]
    risks_sheet_ref = _quote_sheet(risks_info["sheet_name"])
    ws.cell(row=contingency_row, column=1, value=t(lang, "summary_contingency"))
    ws.cell(row=contingency_row, column=2,
            value=f"={risks_sheet_ref}!L{risks_contingency_row}")
    current += 1

    # Fascia BASSA
    bassa_row = current
    ws.cell(row=bassa_row, column=1, value=t(lang, "fascia_bassa"))
    ws.cell(row=bassa_row, column=2, value=f"=B{subtotal_row}+B{contingency_row}")
    current += 1

    # Management Reserve (on Fascia BASSA base)
    mr_row = current
    ws.cell(row=mr_row, column=1,
            value=t(lang, "summary_management_reserve", pct=_pct_label(mr_pct)))
    ws.cell(row=mr_row, column=2, value=f"=B{bassa_row}*{mr_pct}")
    current += 1

    # Fascia MEDIA = BASSA + MR
    media_row = current
    ws.cell(row=media_row, column=1, value=t(lang, "fascia_media"))
    ws.cell(row=media_row, column=2, value=f"=B{bassa_row}+B{mr_row}")
    current += 1

    # Fascia ALTA = MEDIA * (1 + alta_uplift_pct)
    alta_row = current
    ws.cell(row=alta_row, column=1, value=t(lang, "fascia_alta"))
    ws.cell(row=alta_row, column=2, value=f"=B{media_row}*(1+{alta_pct})")
    current += 1

    # Total Billable Effort + ratio (preserved utility from legacy)
    wbs_total_row = wbs_info["total_row"]
    billable_row = current
    ws.cell(row=billable_row, column=1, value=t(lang, "summary_total_billable"))
    ws.cell(row=billable_row, column=2, value=f"=WBS!S{wbs_total_row}")
    current += 1

    ratio_row = current
    ws.cell(row=ratio_row, column=1, value=t(lang, "summary_billable_ratio"))
    ws.cell(row=ratio_row, column=2, value=f"=B{billable_row}/B{tech_row}")
    current += 1

    # --- 3. Calendar Duration ---
    current += 1  # blank separator
    calendar_weeks = _resolve_calendar_weeks(config, phases, rp_info)
    cal_row = current
    ws.cell(row=cal_row, column=1, value=t(lang, "summary_calendar_duration"))
    ws.cell(row=cal_row, column=2, value=calendar_weeks)
    ws.cell(row=cal_row, column=2).number_format = "0"
    current += 1

    # --- 4. Effort by Team (literal PD from WBS primary role) ---
    current += 1  # blank separator
    team_header_row = current
    ws.cell(row=team_header_row, column=1, value=t(lang, "summary_effort_by_team"))
    ws.cell(row=team_header_row, column=1).font = Font(bold=True)
    current += 1
    team_pd = _team_effort_pd(phases, roles)
    for team_name in sorted(team_pd.keys()):
        ws.cell(row=current, column=1, value=team_name)
        # Preserve full precision; Excel renders via number_format
        ws.cell(row=current, column=2, value=team_pd[team_name])
        ws.cell(row=current, column=2).number_format = NUMBER_FMT
        current += 1

    # --- 5. Sensitivity Scenarios ---
    scenarios = data.get("scenarios", []) or []
    if scenarios:
        current += 1
        ws.cell(row=current, column=1, value=t(lang, "summary_sensitivity")).font = Font(bold=True)
        current += 1
        for s in scenarios:
            ws.cell(row=current, column=1, value=str(s))
            current += 1

    _apply_phase_table_formatting(ws, primary_color, data_start, data_end, total_row)
    apply_column_widths(ws, _COLUMN_WIDTHS)


# ---------------------------------------------------------------------------
# Internals
# ---------------------------------------------------------------------------

def _pct_label(pct: float) -> int | float:
    """Convert a 0..1 ratio into a human percentage for label formatting."""
    return int(round(pct * 100)) if pct else 0


def _quote_sheet(name: str) -> str:
    """Quote an Excel sheet reference if it contains spaces or special chars."""
    if any(c in name for c in " '!"):
        return "'" + name.replace("'", "''") + "'"
    return name


def _write_header(ws: Worksheet, effort_unit: str, duration_unit: str) -> None:
    headers = [
        "Phase",
        "Description",
        f"Best Effort ({effort_unit})",
        f"Likely Effort ({effort_unit})",
        f"Worst Effort ({effort_unit})",
        f"PERT Effort ({effort_unit})",
        f"Best Duration ({duration_unit})",
        f"Likely Duration ({duration_unit})",
        f"Worst Duration ({duration_unit})",
        f"PERT Duration ({duration_unit})",
        "σ Duration",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=_HEADER_ROW, column=col_idx, value=header)
        cell.font = HEADER_FONT


def _write_phase_row(ws: Worksheet, summary_row: int, phase: dict, wbs_phase_row: int) -> None:
    ws.cell(row=summary_row, column=_COL_PHASE, value=f"=WBS!B{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_DESC, value=phase.get("description", ""))
    ws.cell(row=summary_row, column=_COL_BEST_E, value=f"=WBS!E{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_LIKELY_E, value=f"=WBS!F{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_WORST_E, value=f"=WBS!G{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_PERT_E, value=f"=WBS!H{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_BEST_D, value=f"=WBS!I{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_LIKELY_D, value=f"=WBS!J{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_WORST_D, value=f"=WBS!K{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_PERT_D, value=f"=WBS!L{wbs_phase_row}")
    ws.cell(row=summary_row, column=_COL_SIGMA, value=f"=WBS!M{wbs_phase_row}")


def _write_total_row(ws: Worksheet, total_row: int, data_start: int, data_end: int) -> None:
    ws.cell(row=total_row, column=_COL_PHASE, value="TOTAL")
    for col_idx in range(_COL_BEST_E, _COL_SIGMA + 1):
        col_letter = get_column_letter(col_idx)
        ws.cell(
            row=total_row,
            column=col_idx,
            value=f"=SUM({col_letter}{data_start}:{col_letter}{data_end})",
        )
    style = get_total_style()
    for col_idx in range(1, _COL_SIGMA + 1):
        apply_style(ws.cell(row=total_row, column=col_idx), style)


def _resolve_calendar_weeks(config: dict, phases: list[dict], rp_info: dict) -> int:
    """Return a single weeks count for the project's calendar duration."""
    explicit = config.get("calendar_total_weeks")
    if explicit:
        return int(explicit)
    starts = [int(p["start_week"]) for p in phases if "start_week" in p]
    ends = [int(p["end_week"]) for p in phases if "end_week" in p]
    if starts and ends:
        return max(ends) - min(starts) + 1
    # Fall back to the Resource Plan total weeks (best effort)
    if rp_info and rp_info.get("total_weeks"):
        return int(rp_info["total_weeks"])
    return 0


def _team_effort_pd(phases: list[dict], roles: list[dict]) -> dict[str, float]:
    role_team = {r["code"]: r.get("team", "Unassigned") for r in roles}
    out: dict[str, float] = {}
    for phase in phases:
        for wp in phase.get("work_packages", []):
            for a in wp.get("activities", []):
                resources = a.get("resources") or []
                if not resources:
                    continue
                primary = resources[0]
                team = role_team.get(primary, "Unassigned")
                pd = (a["best_effort"] + 4 * a["likely_effort"] + a["worst_effort"]) / 6.0
                out[team] = out.get(team, 0.0) + pd
    return out


def _apply_phase_table_formatting(
    ws: Worksheet,
    primary_color: str,
    data_start: int,
    data_end: int,
    total_row: int,
) -> None:
    phase_style = get_phase_style(primary_color)
    formula_fill = get_formula_fill()

    for r in range(data_start, data_end + 1):
        for col_idx in range(1, _COL_SIGMA + 1):
            apply_style(ws.cell(row=r, column=col_idx), phase_style)

    for r in range(data_start, total_row + 1):
        for col_idx in range(_COL_BEST_E, _COL_SIGMA + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.fill = formula_fill
            cell.number_format = NUMBER_FMT
