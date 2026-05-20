"""Integration tests for generate_excel.py orchestrator."""
import importlib.util
import json
import subprocess
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

SCRIPTS_DIR = Path(__file__).parent.parent.resolve()


def _load_generate_excel():
    """Dynamically import generate_excel from the scripts directory."""
    spec = importlib.util.spec_from_file_location(
        "generate_excel",
        SCRIPTS_DIR / "generate_excel.py",
    )
    mod = importlib.util.module_from_spec(spec)
    # Ensure helpers package is importable
    if str(SCRIPTS_DIR) not in sys.path:
        sys.path.insert(0, str(SCRIPTS_DIR))
    spec.loader.exec_module(mod)
    return mod


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def gen():
    """Return the generate_excel module."""
    return _load_generate_excel()


# ---------------------------------------------------------------------------
# Test 1: generates a .xlsx file
# ---------------------------------------------------------------------------

def test_generates_xlsx_file(gen, input_json_path, tmp_xlsx):
    rc = gen.main(str(input_json_path), str(tmp_xlsx))
    assert rc == 0
    assert tmp_xlsx.exists()
    assert tmp_xlsx.suffix == ".xlsx"


# ---------------------------------------------------------------------------
# Test 2: output workbook has exactly 4 sheets in the canonical order
# (English default labels; Italian = "Pianificazione Risorse" / "Riepilogo")
# ---------------------------------------------------------------------------

def test_four_sheets_present_en(gen, input_json_path, tmp_xlsx):
    gen.main(str(input_json_path), str(tmp_xlsx))
    wb = load_workbook(tmp_xlsx)
    assert wb.sheetnames == ["WBS", "Resource Plan", "Risks", "Summary"]
    assert len(wb.sheetnames) == 4


def test_four_sheets_present_it(gen, full_input_data, tmp_path):
    full_input_data["config"]["lang"] = "it"
    json_path = tmp_path / "it.json"
    json_path.write_text(json.dumps(full_input_data))
    out = tmp_path / "out.xlsx"
    gen.main(str(json_path), str(out))
    wb = load_workbook(out)
    assert wb.sheetnames == ["WBS", "Pianificazione Risorse", "Rischi", "Riepilogo"]


def test_no_legacy_sheets(gen, input_json_path, tmp_xlsx):
    """Timeline / Resources must not appear as default sheets anymore."""
    gen.main(str(input_json_path), str(tmp_xlsx))
    wb = load_workbook(tmp_xlsx)
    assert "Timeline" not in wb.sheetnames
    assert "Resources" not in wb.sheetnames


# ---------------------------------------------------------------------------
# Test 3: WBS sheet row count
# Fixture has 2 phases × 1 WP × 2 activities = 1 header + 2 phases + 2 WPs + 4 leaves + 1 TOTAL = 10 rows
# ---------------------------------------------------------------------------

def test_wbs_sheet_has_correct_row_count(gen, input_json_path, tmp_xlsx):
    gen.main(str(input_json_path), str(tmp_xlsx))
    wb = load_workbook(tmp_xlsx)
    ws = wb["WBS"]
    # Count non-empty rows (max_row gives the last row with data)
    assert ws.max_row == 10  # 1 header + 2 phases + 2 WPs + 4 leaves + 1 TOTAL


# ---------------------------------------------------------------------------
# Test 4: formulas intact — WBS!H2 is a formula string
# H = PERT Effort column; row 2 is the first phase row (phase rows have formulas)
# ---------------------------------------------------------------------------

def test_end_to_end_formulas_intact(gen, input_json_path, tmp_xlsx):
    gen.main(str(input_json_path), str(tmp_xlsx))
    wb = load_workbook(tmp_xlsx, data_only=False)
    ws = wb["WBS"]
    cell_value = ws["H2"].value
    assert isinstance(cell_value, str), f"Expected formula string, got {type(cell_value)}: {cell_value!r}"
    assert cell_value.startswith("="), f"Expected formula starting with '=', got: {cell_value!r}"


# ---------------------------------------------------------------------------
# Test 5: cross-sheet references
# Summary should reference WBS; Timeline should reference WBS
# ---------------------------------------------------------------------------

def test_cross_sheet_references(gen, input_json_path, tmp_xlsx):
    gen.main(str(input_json_path), str(tmp_xlsx))
    wb = load_workbook(tmp_xlsx, data_only=False)

    # Summary references WBS
    summary_ws = wb["Summary"]
    summary_formulas = [
        summary_ws.cell(row=r, column=c).value
        for r in range(1, summary_ws.max_row + 1)
        for c in range(1, summary_ws.max_column + 1)
        if isinstance(summary_ws.cell(row=r, column=c).value, str)
        and "WBS!" in summary_ws.cell(row=r, column=c).value
    ]
    assert len(summary_formulas) > 0, "Summary sheet has no WBS cross-references"

    # Risks references WBS (Management Reserve uses WBS!H{total})
    risks_ws = wb["Risks"]
    risks_formulas = [
        risks_ws.cell(row=r, column=c).value
        for r in range(1, risks_ws.max_row + 1)
        for c in range(1, risks_ws.max_column + 1)
        if isinstance(risks_ws.cell(row=r, column=c).value, str)
        and "WBS!" in risks_ws.cell(row=r, column=c).value
    ]
    assert len(risks_formulas) > 0, "Risks sheet has no WBS cross-references"


# ---------------------------------------------------------------------------
# Test 6: CLI interface — exit 0 with valid args
# ---------------------------------------------------------------------------

def test_cli_interface(input_json_path, tmp_xlsx):
    result = subprocess.run(
        [
            sys.executable,
            "generate_excel.py",
            "--input", str(input_json_path),
            "--output", str(tmp_xlsx),
        ],
        cwd=str(SCRIPTS_DIR),
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0, (
        f"CLI exited with {result.returncode}\n"
        f"stdout: {result.stdout}\n"
        f"stderr: {result.stderr}"
    )
    assert tmp_xlsx.exists()


# ---------------------------------------------------------------------------
# Test 7: missing input file → exit non-zero
# ---------------------------------------------------------------------------

def test_missing_input_file_exits_nonzero(tmp_path):
    nonexistent = tmp_path / "does_not_exist.json"
    out = tmp_path / "out.xlsx"
    result = subprocess.run(
        [
            sys.executable,
            "generate_excel.py",
            "--input", str(nonexistent),
            "--output", str(out),
        ],
        cwd=str(SCRIPTS_DIR),
        capture_output=True,
        text=True,
    )
    assert result.returncode != 0


# ---------------------------------------------------------------------------
# Test 8: malformed JSON (missing required keys) → exit 1
# ---------------------------------------------------------------------------

def test_malformed_json_exits_nonzero(tmp_path):
    bad_json = tmp_path / "bad.json"
    # JSON is valid syntax but missing required keys (config, phases, roles, risks)
    bad_json.write_text(json.dumps({"foo": "bar"}))
    out = tmp_path / "out.xlsx"
    result = subprocess.run(
        [
            sys.executable,
            "generate_excel.py",
            "--input", str(bad_json),
            "--output", str(out),
        ],
        cwd=str(SCRIPTS_DIR),
        capture_output=True,
        text=True,
    )
    assert result.returncode != 0


# ---------------------------------------------------------------------------
# Test 9: empty phases array → exit 1
# ---------------------------------------------------------------------------

def test_empty_phases_exits_nonzero(tmp_path, full_input_data):
    empty_phases_data = {**full_input_data, "phases": []}
    json_path = tmp_path / "empty-phases.json"
    json_path.write_text(json.dumps(empty_phases_data))
    out = tmp_path / "out.xlsx"
    result = subprocess.run(
        [
            sys.executable,
            "generate_excel.py",
            "--input", str(json_path),
            "--output", str(out),
        ],
        cwd=str(SCRIPTS_DIR),
        capture_output=True,
        text=True,
    )
    assert result.returncode != 0
