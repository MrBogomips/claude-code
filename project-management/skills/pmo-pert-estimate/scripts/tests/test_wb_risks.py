"""Tests for helpers/wb_risks.py — written FIRST (TDD red phase)."""
import copy

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _import_wb_risks():
    """Import wb_risks module (will fail in RED phase)."""
    from helpers import wb_risks
    return wb_risks


def _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases):
    """Create a fresh Workbook and data dict from fixtures."""
    wb = Workbook()
    data = {
        "config": sample_config,
        "risks": sample_risks,
        "roles": sample_roles,
        "phases": sample_phases,
    }
    return wb, data


# ---------------------------------------------------------------------------
# test_build_creates_sheet
# ---------------------------------------------------------------------------

class TestBuildCreatesSheet:
    def test_build_creates_sheet(self, sample_config, sample_risks, sample_roles, sample_phases):
        """build() must create a sheet named 'Risks' in the workbook."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)
        assert "Risks" in wb.sheetnames


# ---------------------------------------------------------------------------
# test_header_row
# ---------------------------------------------------------------------------

class TestHeaderRow:
    def test_header_row(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Row 1 must contain exactly 13 header labels in columns A–M.

        The Contingency column label must use config.effort_unit for the unit.
        """
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        effort_unit = sample_config["effort_unit"]  # "pd"

        expected_headers = [
            "ID",
            "Risk Description",
            "Category",
            "Affected Phases",
            "Probability (1-5)",
            "Impact (1-5)",
            "Risk Score",
            "Priority",
            "Strategy",
            "Mitigation Action",
            "Owner",
            f"Contingency ({effort_unit})",
            "Contingency Cost",
        ]

        actual_headers = [ws.cell(row=1, column=col).value for col in range(1, 14)]
        assert actual_headers == expected_headers

    def test_header_count_is_13(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Exactly 13 columns — no more, no fewer."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        # Column N (14) must be empty
        assert ws.cell(row=1, column=14).value is None


# ---------------------------------------------------------------------------
# test_risk_score_formula
# ---------------------------------------------------------------------------

class TestRiskScoreFormula:
    def test_risk_score_formula(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Column G (Risk Score) must contain =E{n}*F{n} for each data row."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        num_risks = len(sample_risks)

        for i, row_num in enumerate(range(2, 2 + num_risks)):
            cell_g = ws.cell(row=row_num, column=7)
            expected_formula = f"=E{row_num}*F{row_num}"
            assert cell_g.value == expected_formula, (
                f"Row {row_num}: expected '{expected_formula}', got '{cell_g.value}'"
            )


# ---------------------------------------------------------------------------
# test_priority_formula
# ---------------------------------------------------------------------------

class TestPriorityFormula:
    def test_priority_formula(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Column H (Priority) must contain the nested IF formula for each data row."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        num_risks = len(sample_risks)

        for row_num in range(2, 2 + num_risks):
            cell_h = ws.cell(row=row_num, column=8)
            g_ref = f"G{row_num}"
            expected_formula = (
                f'=IF({g_ref}>=15,"CRITICAL",'
                f'IF({g_ref}>=10,"HIGH",'
                f'IF({g_ref}>=5,"MEDIUM","LOW")))'
            )
            assert cell_h.value == expected_formula, (
                f"Row {row_num}: expected '{expected_formula}', got '{cell_h.value}'"
            )


# ---------------------------------------------------------------------------
# test_contingency_cost_formula
# ---------------------------------------------------------------------------

class TestContingencyCostFormula:
    def test_contingency_cost_formula_with_avg_rate(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """Column M must contain =L{n}*{avg_rate} when avg_rate is set."""
        mod = _import_wb_risks()
        avg_rate = sample_config["avg_rate"]  # 500
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        num_risks = len(sample_risks)

        for row_num in range(2, 2 + num_risks):
            cell_m = ws.cell(row=row_num, column=13)
            expected_formula = f"=L{row_num}*{avg_rate}"
            assert cell_m.value == expected_formula, (
                f"Row {row_num}: expected '{expected_formula}', got '{cell_m.value}'"
            )

    def test_contingency_cost_empty_when_no_avg_rate(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """Column M cells must be empty when avg_rate is None."""
        mod = _import_wb_risks()
        # Create a modified config without avg_rate
        config_no_rate = {**sample_config, "avg_rate": None}
        wb = Workbook()
        data = {
            "config": config_no_rate,
            "risks": sample_risks,
            "roles": sample_roles,
            "phases": sample_phases,
        }
        mod.build(wb, data)

        ws = wb["Risks"]
        num_risks = len(sample_risks)

        for row_num in range(2, 2 + num_risks):
            cell_m = ws.cell(row=row_num, column=13)
            assert cell_m.value is None, (
                f"Row {row_num}: expected None (no avg_rate), got '{cell_m.value}'"
            )


# ---------------------------------------------------------------------------
# test_total_contingency
# ---------------------------------------------------------------------------

class TestTotalContingency:
    def test_total_contingency_formula(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Footer TOTAL row must have =SUM(L{start}:L{end}) in column L."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        sheet_info = mod.build(wb, data)

        ws = wb["Risks"]
        total_row = sheet_info["total_contingency_row"]
        data_start = sheet_info["data_start_row"]
        data_end = sheet_info["data_end_row"]

        cell_l = ws.cell(row=total_row, column=12)
        expected_formula = f"=SUM(L{data_start}:L{data_end})"
        assert cell_l.value == expected_formula, (
            f"Total row L{total_row}: expected '{expected_formula}', got '{cell_l.value}'"
        )

    def test_total_row_comes_after_data(self, sample_config, sample_risks, sample_roles, sample_phases):
        """total_contingency_row must be strictly greater than data_end_row."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        sheet_info = mod.build(wb, data)

        assert sheet_info["total_contingency_row"] > sheet_info["data_end_row"]


# ---------------------------------------------------------------------------
# test_management_reserve
# ---------------------------------------------------------------------------

class TestManagementReserve:
    def test_management_reserve_formula_uses_correct_base(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """MR formula must apply to Tech+Overhead+Contingency, not Contingency alone.

        Form: =(WBS!H{wbs_total}*(1+pm_pct+devops_pct)+L{total_row})*mr_pct
        """
        from helpers import wb_wbs
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        # WBS must exist for the cross-ref to land; pass wbs_info to risks.
        wbs_info = wb_wbs.build(wb, data)
        # Inject modern overhead pcts
        data["config"]["pm_overhead_pct"] = 0.10
        data["config"]["devops_overhead_pct"] = 0.05
        sheet_info = mod.build(wb, data, wbs_info)

        ws = wb["Risks"]
        reserve_row = sheet_info["reserve_row"]
        total_row = sheet_info["total_contingency_row"]
        wbs_total_row = wbs_info["total_row"]
        mr_pct = sample_config["management_reserve_pct"]

        cell_l = ws.cell(row=reserve_row, column=12)
        expected = (
            f"=(WBS!H{wbs_total_row}*(1+0.1+0.05)+L{total_row})*{mr_pct}"
        )
        assert cell_l.value == expected, (
            f"MR formula must use Tech+Overhead+Contingency base. "
            f"Got: {cell_l.value!r}, expected: {expected!r}"
        )

    def test_management_reserve_defaults_to_no_overhead(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """With no overhead config, formula degenerates to (Tech+Contingency)*mr_pct."""
        from helpers import wb_wbs
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        wbs_info = wb_wbs.build(wb, data)
        sheet_info = mod.build(wb, data, wbs_info)

        ws = wb["Risks"]
        cell_l = ws.cell(row=sheet_info["reserve_row"], column=12)
        # No overhead -> multiplier reduces to (1+0+0) (0 or 0.0 both fine).
        value = str(cell_l.value)
        assert "WBS!H" in value
        assert "*(1+0" in value and "+0" in value.split("*(1+0")[1]

    def test_reserve_row_comes_after_total(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """reserve_row must be strictly greater than total_contingency_row."""
        from helpers import wb_wbs
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        wbs_info = wb_wbs.build(wb, data)
        sheet_info = mod.build(wb, data, wbs_info)

        assert sheet_info["reserve_row"] > sheet_info["total_contingency_row"]

    def test_reserve_row_has_label(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """Reserve row column A must have a non-empty label."""
        from helpers import wb_wbs
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        wbs_info = wb_wbs.build(wb, data)
        sheet_info = mod.build(wb, data, wbs_info)

        ws = wb["Risks"]
        reserve_row = sheet_info["reserve_row"]
        label = ws.cell(row=reserve_row, column=1).value
        assert label is not None and label != ""


# ---------------------------------------------------------------------------
# test_input_values
# ---------------------------------------------------------------------------

class TestInputValues:
    def test_probability_is_numeric(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Column E (Probability) must contain numeric values, not formulas."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        for i, risk in enumerate(sample_risks):
            row_num = 2 + i
            cell_e = ws.cell(row=row_num, column=5)
            assert cell_e.value == risk["probability"], (
                f"Row {row_num}: expected probability={risk['probability']}, got '{cell_e.value}'"
            )
            assert isinstance(cell_e.value, (int, float)), (
                f"Row {row_num}: Probability must be numeric, got {type(cell_e.value)}"
            )

    def test_impact_is_numeric(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Column F (Impact) must contain numeric values, not formulas."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        for i, risk in enumerate(sample_risks):
            row_num = 2 + i
            cell_f = ws.cell(row=row_num, column=6)
            assert cell_f.value == risk["impact"], (
                f"Row {row_num}: expected impact={risk['impact']}, got '{cell_f.value}'"
            )
            assert isinstance(cell_f.value, (int, float)), (
                f"Row {row_num}: Impact must be numeric, got {type(cell_f.value)}"
            )

    def test_risk_id_in_column_a(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Column A must contain the risk ID."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        for i, risk in enumerate(sample_risks):
            row_num = 2 + i
            assert ws.cell(row=row_num, column=1).value == risk["id"]

    def test_affected_phases_comma_joined(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """Column D (Affected Phases) must be comma-joined string of phase IDs."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        # R1 has affected_phases = ["1", "2"]
        cell_d2 = ws.cell(row=2, column=4)
        assert cell_d2.value == "1, 2"


# ---------------------------------------------------------------------------
# test_formatting
# ---------------------------------------------------------------------------

class TestFormatting:
    def test_header_bold(self, sample_config, sample_risks, sample_roles, sample_phases):
        """Row 1 header cells must have bold font."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        for col in range(1, 14):
            cell = ws.cell(row=1, column=col)
            assert cell.font is not None and cell.font.bold is True, (
                f"Header cell column {col} must be bold"
            )

    def test_high_score_risks_get_red_font(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """Rows where P*I >= 15 must have red font applied to cells."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        for i, risk in enumerate(sample_risks):
            row_num = 2 + i
            score = risk["probability"] * risk["impact"]
            if score >= 15:
                # At minimum, the score cell (G) should have red font
                cell_g = ws.cell(row=row_num, column=7)
                assert cell_g.font is not None, f"Row {row_num}: score cell should have font set"
                # Red font: color ends with FF0000 or similar red
                rgb = str(cell_g.font.color.rgb).upper() if cell_g.font.color else ""
                assert rgb.endswith("FF0000") or "FF" in rgb[:2], (
                    f"Row {row_num} (score={score}>=15): expected red font, got rgb={rgb}"
                )

    def test_low_score_risks_not_red(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """Rows where P*I < 15 must NOT have red font on score cell."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        mod.build(wb, data)

        ws = wb["Risks"]
        for i, risk in enumerate(sample_risks):
            row_num = 2 + i
            score = risk["probability"] * risk["impact"]
            if score < 15:
                cell_g = ws.cell(row=row_num, column=7)
                if cell_g.font and cell_g.font.color:
                    try:
                        rgb = str(cell_g.font.color.rgb).upper()
                    except Exception:
                        # Color not explicitly set — not red by definition
                        continue
                    # Should not be red (FF0000)
                    assert not rgb.endswith("FF0000"), (
                        f"Row {row_num} (score={score}<15): should NOT have red font, got rgb={rgb}"
                    )

    def test_return_value_has_expected_keys(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """build() must return a dict with the required sheet_info keys."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        sheet_info = mod.build(wb, data)

        assert isinstance(sheet_info, dict)
        required_keys = {
            "sheet_name",
            "data_start_row",
            "data_end_row",
            "total_contingency_row",
            "reserve_row",
        }
        for key in required_keys:
            assert key in sheet_info, f"sheet_info missing key: '{key}'"

    def test_sheet_name_in_return_value(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """sheet_info['sheet_name'] must be 'Risks'."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        sheet_info = mod.build(wb, data)

        assert sheet_info["sheet_name"] == "Risks"

    def test_data_start_row_is_2(self, sample_config, sample_risks, sample_roles, sample_phases):
        """data_start_row must be 2 (row 1 is header)."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        sheet_info = mod.build(wb, data)

        assert sheet_info["data_start_row"] == 2

    def test_data_end_row_matches_risk_count(
        self, sample_config, sample_risks, sample_roles, sample_phases
    ):
        """data_end_row must equal data_start_row + len(risks) - 1."""
        mod = _import_wb_risks()
        wb, data = _build_wb_and_data(sample_config, sample_risks, sample_roles, sample_phases)
        sheet_info = mod.build(wb, data)

        expected_end = sheet_info["data_start_row"] + len(sample_risks) - 1
        assert sheet_info["data_end_row"] == expected_end
