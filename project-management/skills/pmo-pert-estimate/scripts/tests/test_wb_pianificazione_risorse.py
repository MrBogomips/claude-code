"""Tests for the Pianificazione Risorse (Resource Plan) sheet."""
from __future__ import annotations

from datetime import date, timedelta

import pytest
from openpyxl import Workbook

from helpers.i18n import t
from helpers.wb_pianificazione_risorse import build


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _leaf(id_: str, name: str, *, best: float, likely: float, worst: float,
          duration_likely: float = 5, primary_role: str = "SD",
          other_roles: list[str] | None = None) -> dict:
    other_roles = other_roles or []
    return {
        "id": id_,
        "name": name,
        "best_effort": best,
        "likely_effort": likely,
        "worst_effort": worst,
        "best_duration": duration_likely,
        "likely_duration": duration_likely,
        "worst_duration": duration_likely,
        "resources": [primary_role, *other_roles],
        "dependencies": [],
        "risks": [],
        "billable": True,
        "notes": "",
    }


def _phase(id_: str, name: str, activities: list[dict],
           start_week: int | None = None,
           end_week: int | None = None) -> dict:
    p = {
        "id": id_,
        "name": name,
        "description": name,
        "best_duration": 5,
        "likely_duration": 5,
        "worst_duration": 5,
        "work_packages": [{"id": f"{id_}.1", "name": "WP", "activities": activities}],
    }
    if start_week is not None:
        p["start_week"] = start_week
    if end_week is not None:
        p["end_week"] = end_week
    return p


def _data(phases: list[dict], roles: list[dict] | None = None,
          calendar_total_weeks: int | None = None,
          project_start_date: str = "2026-04-06") -> dict:
    roles = roles or [
        {"code": "PM", "name": "Project Manager", "team": "Mgmt", "billable": True},
        {"code": "SD", "name": "Senior Developer", "team": "Dev", "billable": True},
    ]
    cfg = {
        "lang": "en",
        "primary_color": "1B4FA5",
        "project_start_date": project_start_date,
        "pm_overhead_pct": 0.0,
        "devops_overhead_pct": 0.0,
        "alta_uplift_pct": 0.12,
        "calendar_total_weeks": calendar_total_weeks,
    }
    return {"config": cfg, "roles": roles, "phases": phases, "risks": []}


def _wbs_info(phases: list[dict]) -> dict:
    """Minimal wbs_info-like dict; the Resource Plan helper uses it for primary
    role aggregation, but it must also work from `data` alone (defensive)."""
    return {"sheet_name": "WBS"}


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_sheet_created_with_localized_name_en() -> None:
    wb = Workbook(); wb.remove(wb.active)
    data = _data([_phase("1", "Analysis", [_leaf("1.1.1", "A", best=2, likely=4, worst=6)],
                          start_week=1, end_week=2)])
    build(wb, data, _wbs_info(data["phases"]))
    assert "Resource Plan" in wb.sheetnames


def test_sheet_created_with_localized_name_it() -> None:
    wb = Workbook(); wb.remove(wb.active)
    data = _data([_phase("1", "Analysis", [_leaf("1.1.1", "A", best=2, likely=4, worst=6)],
                          start_week=1, end_week=2)])
    data["config"]["lang"] = "it"
    build(wb, data, _wbs_info(data["phases"]))
    assert "Pianificazione Risorse" in wb.sheetnames


def test_one_row_per_active_role() -> None:
    """A role appears in the matrix only if it is the primary role of >= 1 activity."""
    wb = Workbook(); wb.remove(wb.active)
    data = _data([
        _phase("1", "Analysis", [
            _leaf("1.1.1", "A", best=2, likely=4, worst=6, primary_role="SD"),
            _leaf("1.1.2", "B", best=1, likely=2, worst=3, primary_role="PM"),
        ], start_week=1, end_week=2),
    ])
    info = build(wb, data, _wbs_info(data["phases"]))
    assert set(info["role_codes"]) == {"SD", "PM"}


def test_distribution_uniform_across_phase_weeks() -> None:
    """A phase of 10 PD over 5 weeks → 2 PD/week per role-week cell."""
    wb = Workbook(); wb.remove(wb.active)
    # PERT = (5 + 4*10 + 15)/6 = 10
    data = _data([
        _phase("1", "Build", [
            _leaf("1.1.1", "Backend", best=5, likely=10, worst=15, primary_role="SD"),
        ], start_week=1, end_week=5),
    ])
    info = build(wb, data, _wbs_info(data["phases"]))
    ws = wb[info["sheet_name"]]
    # SD row, weeks 1..5: each cell should be 2.0 PD
    sd_row = info["row_by_role"]["SD"]
    week_col_start = info["week_col_start"]
    for w_idx in range(5):
        cell_value = ws.cell(row=sd_row, column=week_col_start + w_idx).value
        assert cell_value == pytest.approx(2.0)


def test_total_pd_per_role_equals_phase_pert_sum() -> None:
    """TOTAL (PD) column per role equals sum of leaf PERT PD where the role is primary."""
    wb = Workbook(); wb.remove(wb.active)
    # SD PERT: leaf1=10 + leaf3=5 = 15
    # PM PERT: leaf2=3
    data = _data([
        _phase("1", "Build", [
            _leaf("1.1.1", "L1", best=5, likely=10, worst=15, primary_role="SD"),
            _leaf("1.1.2", "L2", best=1, likely=3, worst=5, primary_role="PM"),
            _leaf("1.1.3", "L3", best=3, likely=5, worst=7, primary_role="SD"),
        ], start_week=1, end_week=5),
    ])
    info = build(wb, data, _wbs_info(data["phases"]))
    ws = wb[info["sheet_name"]]
    total_col = info["total_col"]
    sd_total = ws.cell(row=info["row_by_role"]["SD"], column=total_col).value
    pm_total = ws.cell(row=info["row_by_role"]["PM"], column=total_col).value
    # TOTAL column is a SUM formula; values not available without computation.
    # Sanity-check via the helper's exposed totals dict instead.
    assert info["role_total_pd"]["SD"] == pytest.approx(15.0)
    assert info["role_total_pd"]["PM"] == pytest.approx(3.0)
    # And the cell holds a SUM formula
    assert isinstance(sd_total, str) and sd_total.startswith("=SUM")
    assert isinstance(pm_total, str) and pm_total.startswith("=SUM")


def test_overlap_phases_sum_per_week() -> None:
    """Two phases overlapping in W3 → that week's role cell is the sum of both contributions."""
    wb = Workbook(); wb.remove(wb.active)
    # Phase1: SD PERT=6 over 3 weeks (W1-W3) => 2/wk
    # Phase2: SD PERT=4 over 2 weeks (W3-W4) => 2/wk
    # W3 SD = 2 + 2 = 4
    data = _data([
        _phase("1", "P1", [_leaf("1.1.1", "x", best=3, likely=6, worst=9, primary_role="SD")],
                start_week=1, end_week=3),
        _phase("2", "P2", [_leaf("2.1.1", "y", best=2, likely=4, worst=6, primary_role="SD")],
                start_week=3, end_week=4),
    ])
    info = build(wb, data, _wbs_info(data["phases"]))
    ws = wb[info["sheet_name"]]
    sd_row = info["row_by_role"]["SD"]
    # W1=2, W2=2, W3=4 (overlap), W4=2
    week_col = info["week_col_start"]
    vals = [ws.cell(row=sd_row, column=week_col + i).value for i in range(4)]
    assert vals[0] == pytest.approx(2.0)
    assert vals[1] == pytest.approx(2.0)
    assert vals[2] == pytest.approx(4.0)
    assert vals[3] == pytest.approx(2.0)


def test_calendar_dates_row() -> None:
    """A calendar row shows project_start_date for W1 and +7 days for each subsequent week."""
    wb = Workbook(); wb.remove(wb.active)
    data = _data(
        [_phase("1", "P1", [_leaf("1.1.1", "x", best=1, likely=1, worst=1)],
                 start_week=1, end_week=3)],
        project_start_date="2026-04-06",
    )
    info = build(wb, data, _wbs_info(data["phases"]))
    ws = wb[info["sheet_name"]]
    cal_row = info["calendar_row"]
    week_col = info["week_col_start"]
    expected = date(2026, 4, 6)
    for i in range(3):
        cell_value = ws.cell(row=cal_row, column=week_col + i).value
        # value can be either a date or an ISO string — accept both
        if isinstance(cell_value, str):
            assert cell_value == expected.isoformat()
        else:
            assert cell_value == expected
        expected += timedelta(days=7)


def test_grand_total_pd_equals_sum_of_role_totals() -> None:
    """Σ all role totals = expected total PD."""
    wb = Workbook(); wb.remove(wb.active)
    data = _data([
        _phase("1", "P1", [
            _leaf("1.1.1", "x", best=5, likely=10, worst=15, primary_role="SD"),
            _leaf("1.1.2", "y", best=1, likely=2, worst=3, primary_role="PM"),
        ], start_week=1, end_week=2),
    ])
    info = build(wb, data, _wbs_info(data["phases"]))
    # Sum of role totals
    grand = sum(info["role_total_pd"].values())
    assert grand == pytest.approx(12.0)  # SD=10 + PM=2


def test_activity_without_primary_role_is_skipped_and_reported() -> None:
    """Activities with empty resources[] are skipped; reported in info['skipped_activities']."""
    wb = Workbook(); wb.remove(wb.active)
    data = _data([
        _phase("1", "P1", [
            {
                "id": "1.1.1", "name": "ghost", "best_effort": 1, "likely_effort": 2,
                "worst_effort": 3, "best_duration": 1, "likely_duration": 1, "worst_duration": 1,
                "resources": [], "dependencies": [], "risks": [], "billable": True, "notes": "",
            },
            _leaf("1.1.2", "real", best=2, likely=4, worst=6, primary_role="SD"),
        ], start_week=1, end_week=2),
    ])
    info = build(wb, data, _wbs_info(data["phases"]))
    assert "1.1.1" in info["skipped_activities"]
    # Only SD present (no phantom role)
    assert set(info["role_codes"]) == {"SD"}


def test_phase_without_explicit_weeks_falls_back_sequentially() -> None:
    """Phases with no start_week/end_week are stacked starting at W1 using duration heuristic."""
    wb = Workbook(); wb.remove(wb.active)
    # 2 phases, no week info; each gets a length derived from PERT duration / 5
    # PERT_d_phase1 = 5 days = 1 week; PERT_d_phase2 = 10 days = 2 weeks
    p1 = _phase("1", "P1", [_leaf("1.1.1", "x", best=1, likely=2, worst=3,
                                   duration_likely=5, primary_role="SD")])
    p2 = _phase("2", "P2", [_leaf("2.1.1", "y", best=1, likely=2, worst=3,
                                   duration_likely=10, primary_role="PM")])
    data = _data([p1, p2])
    info = build(wb, data, _wbs_info(data["phases"]))
    # Phase 1 occupies W1, phase 2 occupies W2-W3 → total_weeks = 3
    assert info["total_weeks"] == 3


def test_weekly_total_row_present_with_localized_label() -> None:
    wb = Workbook(); wb.remove(wb.active)
    data = _data([_phase("1", "P1", [_leaf("1.1.1", "x", best=1, likely=2, worst=3)],
                          start_week=1, end_week=2)])
    info = build(wb, data, _wbs_info(data["phases"]))
    ws = wb[info["sheet_name"]]
    label_cell = ws.cell(row=info["weekly_total_row"], column=1).value
    assert label_cell == t("en", "rp_week_total")


def test_calendar_total_weeks_extends_matrix() -> None:
    """If config.calendar_total_weeks > computed phase span, the matrix is padded with empty weeks."""
    wb = Workbook(); wb.remove(wb.active)
    data = _data(
        [_phase("1", "P1", [_leaf("1.1.1", "x", best=1, likely=2, worst=3)],
                 start_week=1, end_week=2)],
        calendar_total_weeks=5,
    )
    info = build(wb, data, _wbs_info(data["phases"]))
    assert info["total_weeks"] == 5
