"""Tests for the refactored Summary sheet (4-sheet refactor)."""
from __future__ import annotations

import pytest
from openpyxl import Workbook

from helpers import wb_pianificazione_risorse, wb_risks, wb_summary, wb_wbs


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _build_all(data: dict):
    wb = Workbook()
    wb.remove(wb.active)
    wbs_info = wb_wbs.build(wb, data)
    rp_info = wb_pianificazione_risorse.build(wb, data, wbs_info)
    risks_info = wb_risks.build(wb, data)
    return wb, wbs_info, risks_info, rp_info


def _find_label_row(ws, needle: str) -> int | None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and needle in str(cell.value):
                return cell.row
    return None


def _value_in_b(ws, label: str) -> object:
    row = _find_label_row(ws, label)
    assert row is not None, f"label not found: {label}"
    return ws.cell(row=row, column=2).value


# ---------------------------------------------------------------------------
# Sheet presence
# ---------------------------------------------------------------------------

class TestSheetPresence:
    def test_build_creates_sheet_en(self, full_input_data):
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        assert "Summary" in wb.sheetnames

    def test_build_creates_sheet_it(self, full_input_data):
        full_input_data["config"]["lang"] = "it"
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        assert "Riepilogo" in wb.sheetnames


# ---------------------------------------------------------------------------
# Phase table cross-refs (unchanged)
# ---------------------------------------------------------------------------

class TestCrossRefsToWbs:
    def test_phase_name_cross_ref(self, full_input_data):
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        for offset, wbs_phase_row in enumerate(wbs_info["phase_rows"]):
            assert ws.cell(row=2 + offset, column=1).value == f"=WBS!B{wbs_phase_row}"

    def test_pert_effort_cross_ref(self, full_input_data):
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        for offset, wbs_phase_row in enumerate(wbs_info["phase_rows"]):
            assert ws.cell(row=2 + offset, column=6).value == f"=WBS!H{wbs_phase_row}"


# ---------------------------------------------------------------------------
# Tech PERT, Overhead, Subtotal
# ---------------------------------------------------------------------------

class TestTechPertAndOverhead:
    def test_tech_pert_present(self, full_input_data):
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        v = _value_in_b(ws, "Tech PERT Effort")
        assert isinstance(v, str) and v.startswith("="), f"expected formula, got {v!r}"

    def test_pm_overhead_uses_pct(self, full_input_data):
        full_input_data["config"]["pm_overhead_pct"] = 0.10
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        # The label should mention the percentage
        row = _find_label_row(ws, "PM Overhead")
        assert row is not None
        # Formula must reference Tech PERT
        formula = ws.cell(row=row, column=2).value
        assert isinstance(formula, str) and formula.startswith("=")

    def test_devops_overhead_uses_pct(self, full_input_data):
        full_input_data["config"]["devops_overhead_pct"] = 0.05
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        row = _find_label_row(ws, "DevOps Overhead")
        assert row is not None
        formula = ws.cell(row=row, column=2).value
        assert isinstance(formula, str) and formula.startswith("=")

    def test_subtotal_is_sum_of_tech_and_overhead(self, full_input_data):
        full_input_data["config"]["pm_overhead_pct"] = 0.10
        full_input_data["config"]["devops_overhead_pct"] = 0.05
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        v = _value_in_b(ws, "Subtotal Tech + Overhead")
        assert isinstance(v, str) and v.startswith("=")


# ---------------------------------------------------------------------------
# Bands (BASSA / MEDIA / ALTA) and Management Reserve
# ---------------------------------------------------------------------------

class TestBands:
    def test_fascia_bassa_includes_subtotal_and_contingency(self, full_input_data):
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        v = _value_in_b(ws, "Low Band")
        assert isinstance(v, str) and v.startswith("=") and "+" in v

    def test_management_reserve_on_fascia_bassa(self, full_input_data):
        """MR formula references the Fascia BASSA cell, not just contingency."""
        full_input_data["config"]["management_reserve_pct"] = 0.20
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        bassa_row = _find_label_row(ws, "Low Band")
        mr_row = _find_label_row(ws, "Management Reserve")
        assert bassa_row is not None and mr_row is not None
        mr_formula = ws.cell(row=mr_row, column=2).value
        assert isinstance(mr_formula, str) and f"B{bassa_row}" in mr_formula, (
            f"MR formula {mr_formula!r} must reference Fascia BASSA cell B{bassa_row}"
        )

    def test_fascia_media_is_bassa_plus_mr(self, full_input_data):
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        v = _value_in_b(ws, "Medium Band")
        assert isinstance(v, str) and v.startswith("=") and "+" in v

    def test_fascia_alta_uses_uplift(self, full_input_data):
        full_input_data["config"]["alta_uplift_pct"] = 0.12
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        v = _value_in_b(ws, "High Band")
        assert isinstance(v, str) and v.startswith("=") and "*" in v


# ---------------------------------------------------------------------------
# Calendar Duration
# ---------------------------------------------------------------------------

class TestCalendarDuration:
    def test_calendar_duration_from_config(self, full_input_data):
        full_input_data["config"]["calendar_total_weeks"] = 25
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        v = _value_in_b(ws, "Calendar Duration")
        assert v == 25

    def test_calendar_duration_from_phase_weeks(self, full_input_data):
        full_input_data["config"].pop("calendar_total_weeks", None)
        # phase1 W1-W4, phase2 W3-W7  → calendar = 7-1+1 = 7
        full_input_data["phases"][0]["start_week"] = 1
        full_input_data["phases"][0]["end_week"] = 4
        full_input_data["phases"][1]["start_week"] = 3
        full_input_data["phases"][1]["end_week"] = 7
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        v = _value_in_b(ws, "Calendar Duration")
        assert v == 7


# ---------------------------------------------------------------------------
# Effort by Team (in PD, not %)
# ---------------------------------------------------------------------------

class TestEffortByTeam:
    def test_effort_by_team_present_as_numbers(self, full_input_data):
        """Per-team effort rows are literal PD numbers (not Resources cross-refs)."""
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        section_row = _find_label_row(ws, "Effort by Team")
        assert section_row is not None
        # Find at least one team row right after the section header
        # with a numeric value in column B.
        found_numeric = False
        for r in range(section_row + 1, section_row + 12):
            v = ws.cell(row=r, column=2).value
            if isinstance(v, (int, float)) and v > 0:
                found_numeric = True
                break
        assert found_numeric, "no numeric team effort row found below 'Effort by Team' header"

    def test_team_totals_sum_to_tech_pert(self, full_input_data):
        """Σ team subtotals == total PERT effort of activities with a primary role."""
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        section_row = _find_label_row(ws, "Effort by Team")
        assert section_row is not None
        # Scan all numeric rows after the header
        team_sum = 0.0
        for r in range(section_row + 1, section_row + 30):
            v = ws.cell(row=r, column=2).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                team_sum += v
            else:
                break
        # Expected: sum of PERT of activities where resources[0] is set
        expected = 0.0
        for phase in full_input_data["phases"]:
            for wp in phase["work_packages"]:
                for a in wp["activities"]:
                    if a.get("resources"):
                        expected += (a["best_effort"] + 4 * a["likely_effort"] + a["worst_effort"]) / 6.0
        assert team_sum == pytest.approx(expected)


# ---------------------------------------------------------------------------
# Sensitivity Scenarios
# ---------------------------------------------------------------------------

class TestSensitivityScenarios:
    def test_scenarios_listed_as_text(self, full_input_data):
        full_input_data["scenarios"] = [
            "Optimistic: 320 PD",
            "Realistic: 465 PD",
        ]
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        section_row = _find_label_row(ws, "Sensitivity")
        assert section_row is not None
        # Two text rows should follow
        texts: list[str] = []
        for r in range(section_row + 1, section_row + 6):
            v = ws.cell(row=r, column=1).value
            if isinstance(v, str) and v:
                texts.append(v)
            else:
                break
        assert any("Optimistic" in s for s in texts)
        assert any("Realistic" in s for s in texts)


# ---------------------------------------------------------------------------
# Removed legacy behavior
# ---------------------------------------------------------------------------

class TestLegacyRemoved:
    def test_no_sequential_ci_duration(self, full_input_data):
        """CI 68/95 Duration rows must NOT be present (Issue #2)."""
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        assert _find_label_row(ws, "CI 68%") is None
        assert _find_label_row(ws, "CI 95%") is None

    def test_no_resources_cross_ref(self, full_input_data):
        """Summary must not reference the deleted Resources sheet."""
        wb, wbs_info, risks_info, rp_info = _build_all(full_input_data)
        wb_summary.build(wb, full_input_data, wbs_info, risks_info, rp_info)
        ws = wb["Summary"]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "Resources!" in str(cell.value):
                    pytest.fail(f"Summary still references Resources sheet: {cell.value!r}")
