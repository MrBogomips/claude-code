"""Tests for wb_resources.build() — Resources sheet generator.

TDD: tests written BEFORE implementation.
"""
import pytest
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases):
    return {
        "config": sample_config,
        "roles": sample_roles,
        "resource_allocation": sample_resource_allocation,
        "phases": sample_phases,
    }


def _build(data):
    """Import and call the function under test."""
    from helpers.wb_resources import build
    wb = Workbook()
    sheet_info = build(wb, data)
    return wb, sheet_info


# ---------------------------------------------------------------------------
# Fixtures aliases (via conftest)
# ---------------------------------------------------------------------------
# sample_config, sample_roles, sample_resource_allocation, sample_phases
# are all provided by conftest.py


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestBuildCreatesSheet:
    def test_build_creates_sheet(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """build() must create a sheet named 'Resources'."""
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        assert "Resources" in wb.sheetnames
        assert sheet_info["sheet_name"] == "Resources"


class TestMetadataRow:
    def test_metadata_row(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """Row 1 must have billable flags ('Y'/'N') in role columns.

        With 2 roles starting at column D:
          - D1 = 'Y' (SD is billable)
          - E1 = 'N' (DEC is not billable)
        """
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        first_col = sheet_info["first_role_col"]
        first_col_idx = ws[f"{first_col}1"].column  # numeric column index

        for i, role in enumerate(sample_roles):
            col_letter = get_column_letter(first_col_idx + i)
            expected = "Y" if role["billable"] else "N"
            actual = ws[f"{col_letter}1"].value
            assert actual == expected, (
                f"Role {role['code']} (billable={role['billable']}): "
                f"expected '{expected}' in {col_letter}1, got '{actual}'"
            )


class TestHeaderRow:
    def test_header_row(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """Row 2 must have: Phase, Description, Team, [role names], TOTAL EFFORT, BILLABLE EFFORT."""
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        assert ws["A2"].value == "Phase"
        assert ws["B2"].value == "Description"
        assert ws["C2"].value == "Team"

        first_col = sheet_info["first_role_col"]
        first_col_idx = ws[f"{first_col}2"].column

        for i, role in enumerate(sample_roles):
            col_letter = get_column_letter(first_col_idx + i)
            assert ws[f"{col_letter}2"].value == role["name"], (
                f"Expected role name '{role['name']}' in {col_letter}2"
            )

        total_col = sheet_info["total_effort_col"]
        billable_col = sheet_info["billable_effort_col"]
        assert ws[f"{total_col}2"].value == "TOTAL EFFORT"
        assert ws[f"{billable_col}2"].value == "BILLABLE EFFORT"


class TestDynamicRoleColumns:
    def test_dynamic_role_columns(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """Number of role columns must equal len(roles).

        With 2 roles, columns D and E should be role columns.
        """
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)

        role_columns = sheet_info["role_columns"]
        assert len(role_columns) == len(sample_roles)

        # With fixed columns A, B, C the first role column is D
        assert sheet_info["first_role_col"] == "D"
        # With 2 roles, last role col is E
        assert sheet_info["last_role_col"] == "E"

        # Verify mapping
        for role in sample_roles:
            assert role["code"] in role_columns, f"Role {role['code']} missing from role_columns"


class TestDataValues:
    def test_data_values(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """Effort values from resource_allocation must be in correct role×phase cells.

        With the sample data:
          - Phase=Analysis, Team=Dev, SD effort=5 → D_row for (Analysis, Dev)
          - Phase=Analysis, Team=Client, DEC effort=2 → E_row for (Analysis, Client)
          - Phase=Implementation, Team=Dev, SD effort=8 → D_row for (Implementation, Dev)
          - Phase=Implementation, Team=Client, DEC effort=3 → E_row for (Implementation, Client)
        """
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        role_columns = sheet_info["role_columns"]
        data_start = sheet_info["data_start_row"]

        # Build a lookup: (phase_id, team) -> row_number
        row_lookup = {}
        for r in range(data_start, sheet_info["data_end_row"] + 1):
            phase_val = ws[f"A{r}"].value
            team_val = ws[f"C{r}"].value
            if phase_val and team_val:
                row_lookup[(phase_val, team_val)] = r

        # Verify each allocation entry
        for alloc in sample_resource_allocation:
            phase_name = alloc["phase_name"]
            team = alloc["team"]
            role_code = alloc["role_code"]
            expected_effort = alloc["effort"]

            row = row_lookup.get((phase_name, team))
            assert row is not None, f"Row not found for phase={phase_name}, team={team}"

            col = role_columns[role_code]
            actual = ws[f"{col}{row}"].value
            assert actual == expected_effort, (
                f"Expected effort {expected_effort} at {col}{row} "
                f"(phase={phase_name}, team={team}, role={role_code}), got {actual}"
            )


class TestTotalEffortFormula:
    def test_total_effort_formula(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """TOTAL EFFORT column must have =SUM(D_n:E_n) for each data row (2 roles)."""
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        total_col = sheet_info["total_effort_col"]
        first_role = sheet_info["first_role_col"]
        last_role = sheet_info["last_role_col"]
        data_start = sheet_info["data_start_row"]
        data_end = sheet_info["data_end_row"]

        for r in range(data_start, data_end + 1):
            formula = ws[f"{total_col}{r}"].value
            expected = f"=SUM({first_role}{r}:{last_role}{r})"
            assert formula == expected, (
                f"Row {r}: expected '{expected}', got '{formula}'"
            )


class TestBillableEffortFormula:
    def test_billable_effort_formula(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """BILLABLE EFFORT column must use SUMPRODUCT with the Y/N metadata row."""
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        billable_col = sheet_info["billable_effort_col"]
        first_role = sheet_info["first_role_col"]
        last_role = sheet_info["last_role_col"]
        data_start = sheet_info["data_start_row"]
        data_end = sheet_info["data_end_row"]

        for r in range(data_start, data_end + 1):
            formula = ws[f"{billable_col}{r}"].value
            expected = (
                f'=SUMPRODUCT({first_role}{r}:{last_role}{r},'
                f'({first_role}$1:{last_role}$1="Y")*1)'
            )
            assert formula == expected, (
                f"Row {r}: expected '{expected}', got '{formula}'"
            )


class TestFooterTotals:
    def test_footer_totals(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """Footer row must contain TOTAL EFFORT sum and TOTAL BILLABLE sum formulas."""
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        total_col = sheet_info["total_effort_col"]
        billable_col = sheet_info["billable_effort_col"]
        data_start = sheet_info["data_start_row"]
        data_end = sheet_info["data_end_row"]
        footer_row = data_end + 1  # footer is the row right after data rows

        # Footer label
        footer_label = ws[f"A{footer_row}"].value
        assert footer_label is not None and str(footer_label).upper() in ("TOTAL", "GRAND TOTAL"), (
            f"Expected TOTAL label in A{footer_row}, got '{footer_label}'"
        )

        # TOTAL EFFORT footer formula: sums team subtotal rows (not full range,
        # to avoid double-counting data rows that are already in subtotals)
        total_formula = ws[f"{total_col}{footer_row}"].value
        assert total_formula is not None and total_formula.startswith("=SUM("), (
            f"Expected SUM formula in {total_col}{footer_row}, got '{total_formula}'"
        )
        # Verify it references team subtotal rows
        team_subtotal_rows = sheet_info["team_subtotal_rows"]
        for team_row in team_subtotal_rows.values():
            assert str(team_row) in total_formula, (
                f"Footer formula '{total_formula}' should reference subtotal row {team_row}"
            )

        # BILLABLE EFFORT footer formula
        billable_formula = ws[f"{billable_col}{footer_row}"].value
        assert billable_formula is not None and billable_formula.startswith("=SUM("), (
            f"Expected SUM formula in {billable_col}{footer_row}, got '{billable_formula}'"
        )


class TestTeamSubtotals:
    def test_team_subtotals(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """One subtotal row per distinct team must exist in team_subtotal_rows."""
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        distinct_teams = {alloc["team"] for alloc in sample_resource_allocation}
        team_subtotal_rows = sheet_info["team_subtotal_rows"]

        assert set(team_subtotal_rows.keys()) == distinct_teams, (
            f"Expected subtotal rows for teams {distinct_teams}, "
            f"got {set(team_subtotal_rows.keys())}"
        )

        # Each subtotal row must have a formula or label
        total_col = sheet_info["total_effort_col"]
        for team, row in team_subtotal_rows.items():
            cell_value = ws[f"{total_col}{row}"].value
            assert cell_value is not None, (
                f"Team '{team}' subtotal row {row}: {total_col}{row} is empty"
            )
            assert str(cell_value).startswith("="), (
                f"Team '{team}' subtotal row {row}: expected formula, got '{cell_value}'"
            )


class TestFormatting:
    def test_formatting(
        self, sample_config, sample_roles, sample_resource_allocation, sample_phases
    ):
        """Header row (row 2) must be bold; metadata row (row 1) italic; data rows normal."""
        data = _make_data(sample_config, sample_roles, sample_resource_allocation, sample_phases)
        wb, sheet_info = _build(data)
        ws = wb["Resources"]

        # Header row (row 2) — cells should be bold
        header_cell = ws["A2"]
        assert header_cell.font is not None and header_cell.font.bold, (
            f"Header cell A2 should be bold, font={header_cell.font}"
        )

        # Metadata row (row 1) — role cells should be italic
        first_role_col = sheet_info["first_role_col"]
        metadata_cell = ws[f"{first_role_col}1"]
        assert metadata_cell.font is not None and metadata_cell.font.italic, (
            f"Metadata cell {first_role_col}1 should be italic, font={metadata_cell.font}"
        )

        # Data rows — first data cell should NOT be bold
        data_start = sheet_info["data_start_row"]
        data_cell = ws[f"A{data_start}"]
        is_bold = data_cell.font is not None and data_cell.font.bold
        assert not is_bold, (
            f"Data cell A{data_start} should not be bold, font={data_cell.font}"
        )
