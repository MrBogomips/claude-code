"""Tests for the Timeline/Gantt sheet generator (helpers/wb_timeline.py).

TDD: these tests are written FIRST and must fail before the implementation exists.

Test setup:
  1. Create a Workbook
  2. Build data dict from fixtures
  3. Call wb_wbs.build(wb, data) first to get wbs_info
  4. Then call wb_timeline.build(wb, data, wbs_info)
"""
import math

import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

from helpers import wb_wbs
from helpers import wb_timeline  # noqa: E402


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_data(sample_config, sample_phases, sample_roles):
    return {
        "config": sample_config,
        "phases": sample_phases,
        "roles": sample_roles,
    }


def _build_both(wb, data):
    """Build WBS then Timeline sheets; return (wbs_info, timeline_info)."""
    wbs_info = wb_wbs.build(wb, data)
    timeline_info = wb_timeline.build(wb, data, wbs_info)
    return wbs_info, timeline_info


def _get_ws(wb):
    return wb["Timeline"]


def _col_letter(n: int) -> str:
    """Convert 1-based column index to letter(s), e.g. 1->'A', 27->'AA'."""
    result = ""
    while n > 0:
        n, rem = divmod(n - 1, 26)
        result = chr(65 + rem) + result
    return result


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestBuildCreatesSheet:
    """wb_timeline.build() must create a sheet named 'Timeline'."""

    def test_build_creates_sheet(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        _build_both(wb, data)
        assert "Timeline" in wb.sheetnames

    def test_build_returns_sheet_info(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info = wb_wbs.build(wb, data)
        result = wb_timeline.build(wb, data, wbs_info)
        assert isinstance(result, dict)
        assert "sheet_name" in result
        assert result["sheet_name"] == "Timeline"
        assert "num_periods" in result
        assert "summary_start_row" in result


class TestCrossReferencesToWbs:
    """Columns A-D in data rows must cross-reference the WBS sheet."""

    def test_cross_references_to_wbs(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        data_start = wbs_info["data_start_row"]
        data_end = wbs_info["data_end_row"]

        # Check first data row
        r = data_start
        val_a = ws[f"A{r}"].value
        assert isinstance(val_a, str) and "WBS" in val_a and f"A{r}" in val_a, \
            f"Col A row {r} should reference WBS!A{r}, got {val_a!r}"

        val_b = ws[f"B{r}"].value
        assert isinstance(val_b, str) and "WBS" in val_b, \
            f"Col B row {r} should have WBS concatenation formula, got {val_b!r}"

        val_c = ws[f"C{r}"].value
        assert isinstance(val_c, str) and "WBS" in val_c and f"L{r}" in val_c, \
            f"Col C row {r} should reference WBS!L{r} (PERT Duration), got {val_c!r}"

        val_d = ws[f"D{r}"].value
        assert isinstance(val_d, str) and "WBS" in val_d and f"M{r}" in val_d, \
            f"Col D row {r} should reference WBS!M{r} (sigma), got {val_d!r}"

    def test_col_b_concatenation_formula(self, sample_config, sample_phases, sample_roles):
        """Col B formula must concatenate WBS B, C, D columns."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, _ = _build_both(wb, data)
        ws = _get_ws(wb)

        r = wbs_info["data_start_row"]
        val_b = ws[f"B{r}"].value
        # Must reference B, C, D columns of WBS
        assert "WBS!B" in val_b or "WBS!C" in val_b or "WBS!D" in val_b, \
            f"Col B formula should reference WBS B/C/D columns: {val_b!r}"

    def test_all_data_rows_have_cross_refs(self, sample_config, sample_phases, sample_roles):
        """All data rows (phases, WPs, leaves) must have cross-reference formulas."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, _ = _build_both(wb, data)
        ws = _get_ws(wb)

        for r in range(wbs_info["data_start_row"], wbs_info["data_end_row"] + 1):
            val_a = ws[f"A{r}"].value
            assert isinstance(val_a, str) and val_a.startswith("="), \
                f"Row {r} Col A should be a formula, got {val_a!r}"


class TestPeriodColumns:
    """Period header columns (P1, P2, ...) must be generated based on config and total duration."""

    def test_period_columns_exist(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        num_periods = timeline_info["num_periods"]
        assert num_periods >= 1, "There should be at least 1 period column"

    def test_period_headers_labeled(self, sample_config, sample_phases, sample_roles):
        """Header row must have P1, P2, ... labels for period columns."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        num_periods = timeline_info["num_periods"]
        # Period columns start at column 5 (E) — after ID, Label, PERT Dur, σ
        period_col_start = 5
        header_row = 1
        for p in range(1, num_periods + 1):
            col_idx = period_col_start + p - 1
            col_letter = _col_letter(col_idx)
            header_val = ws[f"{col_letter}{header_row}"].value
            assert header_val == f"P{p}", \
                f"Expected 'P{p}' in column {col_letter}, got {header_val!r}"

    def test_biweekly_period_type(self, sample_config, sample_phases, sample_roles):
        """With biweekly config, periods should be 10-working-day blocks."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)

        # With sample data, PERT durations total several days.
        # Biweekly = 10 working days per period.
        # There should be at least 1 period.
        assert timeline_info["num_periods"] >= 1

    def test_period_count_matches_duration(self, sample_config, sample_phases, sample_roles):
        """num_periods == ceil(total_duration / period_length)."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)

        # Compute expected PERT durations manually:
        # 1.1.1: (2+4*3+5)/6 = 19/6 ≈ 3.167
        # 1.1.2: (1+4*2+3)/6 = 12/6 = 2.0  (depends on 1.1.1)
        # 2.1.1: (5+4*7+10)/6 = 43/6 ≈ 7.167  (depends on 1.1.2)
        # 2.1.2: (10+4*14+20)/6 = 86/6 ≈ 14.33  (no deps → starts at 0)
        # Critical path: 1.1.1 → 1.1.2 → 2.1.1
        # total_duration = ES(2.1.1) + dur(2.1.1)
        # ES(1.1.1)=0, EF(1.1.1)=3.167
        # ES(1.1.2)=3.167, EF(1.1.2)=5.167
        # ES(2.1.1)=5.167, EF(2.1.1)=12.333
        # total = max(12.333, 14.333) = 14.333
        total_dur = max(
            19/6 + 12/6 + 43/6,   # chain end
            86/6                   # 2.1.2 standalone
        )
        period_length = 10  # biweekly
        expected_periods = math.ceil(total_dur / period_length)

        wbs_info = wb_wbs.build(wb, data)
        timeline_info = wb_timeline.build(wb, data, wbs_info)

        assert timeline_info["num_periods"] == expected_periods, \
            f"Expected {expected_periods} periods, got {timeline_info['num_periods']}"


class TestCriticalPathColoring:
    """Activities on the critical path get red (FF0000) fill in their Gantt cells."""

    def _find_gantt_fill_for_activity(self, ws, activity_id, wbs_ws, wbs_info, timeline_info):
        """Find the Gantt cells for an activity row and return their fill colors."""
        # Find the WBS row for this activity
        wbs_row = None
        for r in range(wbs_info["data_start_row"], wbs_info["data_end_row"] + 1):
            if wbs_ws[f"A{r}"].value == activity_id:
                wbs_row = r
                break
        if wbs_row is None:
            return []

        # The Timeline row is the same number (1-to-1 mapping)
        tl_row = wbs_row
        period_col_start = 5
        num_periods = timeline_info["num_periods"]

        fills = []
        for p in range(num_periods):
            col_letter = _col_letter(period_col_start + p)
            cell = ws[f"{col_letter}{tl_row}"]
            if cell.fill and cell.fill.fill_type == "solid":
                fills.append(cell.fill.fgColor.rgb)
            else:
                fills.append(None)
        return fills

    def test_critical_path_coloring(self, sample_config, sample_phases, sample_roles):
        """Activities on critical path have red Gantt cells."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)
        wbs_ws = wb["WBS"]

        # In sample data, 2.1.1 depends on 1.1.2 which depends on 1.1.1 → critical path
        fills = self._find_gantt_fill_for_activity(
            ws, "2.1.1", wbs_ws, wbs_info, timeline_info
        )
        # At least one Gantt cell should be red
        red_cells = [f for f in fills if f is not None and "FF0000" in f]
        assert len(red_cells) >= 1, \
            f"Expected red fill for critical path activity 2.1.1, fills={fills}"

    def test_critical_path_start_activity_colored(self, sample_config, sample_phases, sample_roles):
        """First activity on critical path (1.1.1) also gets red fill."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)
        wbs_ws = wb["WBS"]

        fills = self._find_gantt_fill_for_activity(
            ws, "1.1.1", wbs_ws, wbs_info, timeline_info
        )
        red_cells = [f for f in fills if f is not None and "FF0000" in f]
        assert len(red_cells) >= 1, \
            f"Expected red fill for critical path activity 1.1.1, fills={fills}"


class TestParallelColoring:
    """Non-critical activities that have dependencies get blue (4472C4) fill."""

    def test_parallel_coloring(self, sample_config, sample_phases, sample_roles):
        """If any activity has deps but is not on critical path, it gets blue fill.

        With the sample data as-is, all dep-chains lead to the critical path.
        We modify the data to create a parallel non-critical activity.
        """
        # Create a scenario where there's a parallel non-critical activity
        import copy
        phases = copy.deepcopy(sample_phases)
        # Add an activity that depends on 1.1.1 but is much shorter → not critical
        phases[0]["work_packages"][0]["activities"].append({
            "id": "1.1.3",
            "name": "Quick review",
            "best_effort": 0.5,
            "likely_effort": 0.5,
            "worst_effort": 1,
            "best_duration": 0.5,
            "likely_duration": 0.5,
            "worst_duration": 1,
            "resources": ["SD"],
            "dependencies": ["1.1.1"],  # has deps but very short → not critical
            "risks": [],
            "billable": True,
            "notes": "",
        })

        wb = Workbook()
        data = {
            "config": sample_config,
            "phases": phases,
            "roles": sample_roles,
        }
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)
        wbs_ws = wb["WBS"]

        # Find 1.1.3 row in WBS
        row_1_1_3 = None
        for r in range(wbs_info["data_start_row"], wbs_info["data_end_row"] + 1):
            if wbs_ws[f"A{r}"].value == "1.1.3":
                row_1_1_3 = r
                break
        assert row_1_1_3 is not None, "Activity 1.1.3 not found in WBS"

        # Check Timeline cells for 1.1.3
        period_col_start = 5
        num_periods = timeline_info["num_periods"]
        blue_cells = []
        for p in range(num_periods):
            col_letter = _col_letter(period_col_start + p)
            cell = ws[f"{col_letter}{row_1_1_3}"]
            if cell.fill and cell.fill.fill_type == "solid":
                rgb = cell.fill.fgColor.rgb
                if "4472C4" in rgb:
                    blue_cells.append(rgb)

        assert len(blue_cells) >= 1, \
            f"Expected blue fill for parallel activity 1.1.3, got no blue cells"


class TestContinuousColoring:
    """Activities with no deps and no dependents get orange (FFA500) fill."""

    def test_continuous_coloring(self, sample_config, sample_phases, sample_roles):
        """2.1.2 has no dependencies and no dependents → continuous (orange)."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)
        wbs_ws = wb["WBS"]

        # Find 2.1.2 in WBS
        row_2_1_2 = None
        for r in range(wbs_info["data_start_row"], wbs_info["data_end_row"] + 1):
            if wbs_ws[f"A{r}"].value == "2.1.2":
                row_2_1_2 = r
                break
        assert row_2_1_2 is not None, "Activity 2.1.2 not found in WBS"

        period_col_start = 5
        num_periods = timeline_info["num_periods"]
        orange_cells = []
        for p in range(num_periods):
            col_letter = _col_letter(period_col_start + p)
            cell = ws[f"{col_letter}{row_2_1_2}"]
            if cell.fill and cell.fill.fill_type == "solid":
                rgb = cell.fill.fgColor.rgb
                if "FFA500" in rgb:
                    orange_cells.append(rgb)

        assert len(orange_cells) >= 1, \
            f"Expected orange fill for continuous activity 2.1.2, got no orange cells"


class TestGanttCellPlacement:
    """Gantt cells must be in the correct period columns based on ES and PERT duration."""

    def _compute_pert(self, b, l, w):
        return (b + 4 * l + w) / 6

    def test_gantt_cell_placement(self, sample_config, sample_phases, sample_roles):
        """Activity 1.1.1 has ES=0 and PERT≈3.17 days.
        With biweekly (10-day) periods, it should be in period 1 only (cols E+).
        """
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)
        wbs_ws = wb["WBS"]

        # Find 1.1.1 row
        row_1_1_1 = None
        for r in range(wbs_info["data_start_row"], wbs_info["data_end_row"] + 1):
            if wbs_ws[f"A{r}"].value == "1.1.1":
                row_1_1_1 = r
                break
        assert row_1_1_1 is not None

        # 1.1.1: ES=0, PERT=(2+4*3+5)/6 ≈ 3.17, period_length=10
        # Should be in period 1 (P1)
        period_col_start = 5
        num_periods = timeline_info["num_periods"]
        filled_periods = []
        for p in range(num_periods):
            col_letter = _col_letter(period_col_start + p)
            cell = ws[f"{col_letter}{row_1_1_1}"]
            if cell.fill and cell.fill.fill_type == "solid" and \
               cell.fill.fgColor.rgb not in ("00000000", "FF000000", None):
                filled_periods.append(p + 1)  # 1-based

        assert 1 in filled_periods, \
            f"Activity 1.1.1 should have a Gantt cell in period 1; filled={filled_periods}"

    def test_gantt_only_for_leaf_activities(self, sample_config, sample_phases, sample_roles):
        """Phase and WP rows must NOT have Gantt cells filled."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)
        wbs_ws = wb["WBS"]

        phase_rows = wbs_info["phase_rows"]
        period_col_start = 5
        num_periods = timeline_info["num_periods"]

        for r in phase_rows:
            for p in range(num_periods):
                col_letter = _col_letter(period_col_start + p)
                cell = ws[f"{col_letter}{r}"]
                # Phase rows must not have colored Gantt fills
                if cell.fill and cell.fill.fill_type == "solid":
                    rgb = cell.fill.fgColor.rgb
                    assert rgb in ("00000000", "FF000000", None, "FFFFFFFF"), \
                        f"Phase row {r} col {col_letter} should not have Gantt fill, got {rgb}"


class TestSummaryBlock:
    """Below Gantt data: rows with key project summary labels."""

    def test_summary_block_exists(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        assert summary_start > wbs_info["data_end_row"], \
            "Summary block must start below data rows"

    def test_summary_labels(self, sample_config, sample_phases, sample_roles):
        """Summary block has required label rows."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        # Collect all label values from col A in the summary section
        labels = []
        r = summary_start
        while True:
            val = ws[f"A{r}"].value
            if val is None:
                break
            labels.append(str(val))
            r += 1

        required_labels = ["Critical Path", "Total PERT Duration", "CI 68%", "CI 95%"]
        for req in required_labels:
            assert any(req in lbl for lbl in labels), \
                f"Summary block missing label '{req}'; found: {labels}"

    def test_summary_ci_formulas(self, sample_config, sample_phases, sample_roles):
        """CI 68% and CI 95% rows must have formula values in col B."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        ci68_row = None
        ci95_row = None
        r = summary_start
        while True:
            val = ws[f"A{r}"].value
            if val is None:
                break
            if "CI 68%" in str(val):
                ci68_row = r
            if "CI 95%" in str(val):
                ci95_row = r
            r += 1

        assert ci68_row is not None, "CI 68% row not found in summary"
        assert ci95_row is not None, "CI 95% row not found in summary"

        # Col B must have formula or numeric value
        ci68_val = ws[f"B{ci68_row}"].value
        ci95_val = ws[f"B{ci95_row}"].value
        assert ci68_val is not None, "CI 68% row col B should have a value or formula"
        assert ci95_val is not None, "CI 95% row col B should have a value or formula"


class TestLegendRow:
    """Legend with 3 color entries (Critical, Parallel, Continuous) below summary."""

    def test_legend_row_exists(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        # Scan rows below summary_start for legend entries
        legend_labels = []
        for r in range(summary_start, summary_start + 30):
            val = ws[f"A{r}"].value
            if val is not None:
                legend_labels.append(str(val))

        # Check for legend markers
        legend_text = " ".join(legend_labels)
        assert any(kw in legend_text for kw in ["Critical", "Parallel", "Continuous", "Legend"]), \
            f"No legend found below summary; labels found: {legend_labels}"

    def test_legend_has_three_color_entries(self, sample_config, sample_phases, sample_roles):
        """Legend section must have 3 distinct colored cells."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        # Scan entire summary + legend area for colored cells
        colored_cells = set()
        for r in range(summary_start, summary_start + 30):
            for c in range(1, 10):
                from openpyxl.utils import get_column_letter
                col = get_column_letter(c)
                cell = ws[f"{col}{r}"]
                if cell.fill and cell.fill.fill_type == "solid":
                    rgb = cell.fill.fgColor.rgb
                    if rgb not in ("00000000", "FF000000", "FFFFFFFF", None):
                        colored_cells.add(rgb)

        assert len(colored_cells) >= 3, \
            f"Expected at least 3 distinct legend colors, found: {colored_cells}"

    def test_legend_colors_are_correct(self, sample_config, sample_phases, sample_roles):
        """Legend cells must use the exact expected colors."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        colored_cells = set()
        for r in range(summary_start, summary_start + 30):
            for c in range(1, 10):
                from openpyxl.utils import get_column_letter
                col = get_column_letter(c)
                cell = ws[f"{col}{r}"]
                if cell.fill and cell.fill.fill_type == "solid":
                    rgb = cell.fill.fgColor.rgb
                    if rgb not in ("00000000", "FF000000", "FFFFFFFF", None):
                        colored_cells.add(rgb)

        # All three legend colors must be present
        expected_colors = {"FFFF0000", "FF4472C4", "FFFFA500"}
        for color in expected_colors:
            assert color in colored_cells, \
                f"Legend color {color} not found; found colors: {colored_cells}"


class TestIndicativeDates:
    """When config.start_date is provided, summary block includes Start/End date rows."""

    def test_indicative_dates_present_when_start_date_set(
        self, sample_config, sample_phases, sample_roles
    ):
        """When start_date is in config, Start Date and End Date appear in summary."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        # sample_config has start_date = "2026-04-06"
        assert data["config"].get("start_date") == "2026-04-06"

        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        labels = []
        for r in range(summary_start, summary_start + 30):
            val = ws[f"A{r}"].value
            if val is not None:
                labels.append(str(val))

        label_text = " ".join(labels)
        assert "Start Date" in label_text, \
            f"Expected 'Start Date' in summary when start_date set; labels={labels}"
        assert "End Date" in label_text, \
            f"Expected 'End Date' in summary when start_date set; labels={labels}"

    def test_indicative_dates_absent_when_no_start_date(
        self, sample_config, sample_phases, sample_roles
    ):
        """When start_date is None or absent, no Start/End date in summary."""
        import copy
        cfg = copy.deepcopy(sample_config)
        cfg.pop("start_date", None)

        wb = Workbook()
        data = {
            "config": cfg,
            "phases": sample_phases,
            "roles": sample_roles,
        }
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        labels = []
        for r in range(summary_start, summary_start + 30):
            val = ws[f"A{r}"].value
            if val is not None:
                labels.append(str(val))

        label_text = " ".join(labels)
        assert "Start Date" not in label_text, \
            f"'Start Date' should not appear when no start_date; labels={labels}"
        assert "End Date" not in label_text, \
            f"'End Date' should not appear when no start_date; labels={labels}"

    def test_end_date_is_calculated(self, sample_config, sample_phases, sample_roles):
        """When start_date is set, End Date row in col B contains a date or formula."""
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wbs_info, timeline_info = _build_both(wb, data)
        ws = _get_ws(wb)

        summary_start = timeline_info["summary_start_row"]
        end_date_val = None
        for r in range(summary_start, summary_start + 30):
            val = ws[f"A{r}"].value
            if val is not None and "End Date" in str(val):
                end_date_val = ws[f"B{r}"].value
                break

        assert end_date_val is not None, \
            "End Date row col B should have a value (date or formula)"
