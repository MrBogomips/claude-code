"""Tests for helpers/formatting.py — written FIRST (TDD red phase)."""
import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _import_formatting():
    """Import the formatting module (may fail in RED phase)."""
    from helpers import formatting
    return formatting


# ---------------------------------------------------------------------------
# get_phase_style
# ---------------------------------------------------------------------------

class TestGetPhaseStyle:
    def test_returns_dict(self):
        fmt = _import_formatting()
        style = fmt.get_phase_style("1B4FA5")
        assert isinstance(style, dict)

    def test_has_font_key(self):
        fmt = _import_formatting()
        style = fmt.get_phase_style("1B4FA5")
        assert "font" in style

    def test_font_is_bold(self):
        fmt = _import_formatting()
        style = fmt.get_phase_style("1B4FA5")
        assert style["font"].bold is True

    def test_font_color_is_white(self):
        fmt = _import_formatting()
        style = fmt.get_phase_style("1B4FA5")
        # openpyxl Color rgb is stored as AARRGGBB
        assert style["font"].color.rgb in ("FFFFFFFF", "FF000000"[:-6] + "FFFFFF")
        # More precisely:
        assert style["font"].color.rgb.upper().endswith("FFFFFF")

    def test_fill_matches_primary_color(self):
        fmt = _import_formatting()
        primary = "1B4FA5"
        style = fmt.get_phase_style(primary)
        assert "fill" in style
        assert style["fill"].fgColor.rgb.upper().endswith(primary.upper())

    def test_different_primary_colors(self):
        fmt = _import_formatting()
        style_a = fmt.get_phase_style("1B4FA5")
        style_b = fmt.get_phase_style("AA0000")
        assert style_a["fill"].fgColor.rgb != style_b["fill"].fgColor.rgb

    def test_immutability_returns_new_objects(self):
        fmt = _import_formatting()
        s1 = fmt.get_phase_style("1B4FA5")
        s2 = fmt.get_phase_style("1B4FA5")
        # Must be distinct dict instances
        assert s1 is not s2
        # And distinct Font instances
        assert s1["font"] is not s2["font"]


# ---------------------------------------------------------------------------
# get_wp_style
# ---------------------------------------------------------------------------

class TestGetWpStyle:
    def test_returns_dict(self):
        fmt = _import_formatting()
        assert isinstance(fmt.get_wp_style(), dict)

    def test_font_is_bold(self):
        fmt = _import_formatting()
        assert fmt.get_wp_style()["font"].bold is True

    def test_fill_color_d9e2f3(self):
        fmt = _import_formatting()
        style = fmt.get_wp_style()
        assert "fill" in style
        assert style["fill"].fgColor.rgb.upper().endswith("D9E2F3")

    def test_immutability(self):
        fmt = _import_formatting()
        s1 = fmt.get_wp_style()
        s2 = fmt.get_wp_style()
        assert s1 is not s2
        assert s1["font"] is not s2["font"]


# ---------------------------------------------------------------------------
# get_leaf_style
# ---------------------------------------------------------------------------

class TestGetLeafStyle:
    def test_returns_dict(self):
        fmt = _import_formatting()
        assert isinstance(fmt.get_leaf_style(), dict)

    def test_font_is_not_bold(self):
        fmt = _import_formatting()
        style = fmt.get_leaf_style()
        assert style["font"].bold is not True

    def test_no_fill_or_white_fill(self):
        fmt = _import_formatting()
        style = fmt.get_leaf_style()
        # Either no fill key, or fill is none/white
        if "fill" in style:
            fill = style["fill"]
            # none patternType or white color
            assert fill.patternType in (None, "none") or fill.fgColor.rgb.upper().endswith("FFFFFF") or fill.fgColor.rgb.upper() == "00000000"

    def test_immutability(self):
        fmt = _import_formatting()
        s1 = fmt.get_leaf_style()
        s2 = fmt.get_leaf_style()
        assert s1 is not s2


# ---------------------------------------------------------------------------
# get_total_style
# ---------------------------------------------------------------------------

class TestGetTotalStyle:
    def test_returns_dict(self):
        fmt = _import_formatting()
        assert isinstance(fmt.get_total_style(), dict)

    def test_font_is_bold(self):
        fmt = _import_formatting()
        assert fmt.get_total_style()["font"].bold is True

    def test_font_color_is_white(self):
        fmt = _import_formatting()
        style = fmt.get_total_style()
        assert style["font"].color.rgb.upper().endswith("FFFFFF")

    def test_fill_is_dark_2f2f2f(self):
        fmt = _import_formatting()
        style = fmt.get_total_style()
        assert "fill" in style
        assert style["fill"].fgColor.rgb.upper().endswith("2F2F2F")

    def test_has_double_top_border(self):
        fmt = _import_formatting()
        style = fmt.get_total_style()
        assert "border" in style
        top = style["border"].top
        assert top is not None
        assert top.border_style == "double"

    def test_immutability(self):
        fmt = _import_formatting()
        s1 = fmt.get_total_style()
        s2 = fmt.get_total_style()
        assert s1 is not s2
        assert s1["font"] is not s2["font"]


# ---------------------------------------------------------------------------
# get_formula_fill
# ---------------------------------------------------------------------------

class TestGetFormulaFill:
    def test_returns_pattern_fill(self):
        fmt = _import_formatting()
        fill = fmt.get_formula_fill()
        assert isinstance(fill, PatternFill)

    def test_fill_color_fff2cc(self):
        fmt = _import_formatting()
        fill = fmt.get_formula_fill()
        assert fill.fgColor.rgb.upper().endswith("FFF2CC")

    def test_immutability(self):
        fmt = _import_formatting()
        f1 = fmt.get_formula_fill()
        f2 = fmt.get_formula_fill()
        assert f1 is not f2


# ---------------------------------------------------------------------------
# get_number_format
# ---------------------------------------------------------------------------

class TestGetNumberFormat:
    def test_returns_correct_string(self):
        fmt = _import_formatting()
        assert fmt.get_number_format() == "#,##0.00"

    def test_number_fmt_constant(self):
        fmt = _import_formatting()
        assert hasattr(fmt, "NUMBER_FMT")
        assert fmt.NUMBER_FMT == "#,##0.00"


# ---------------------------------------------------------------------------
# apply_column_widths
# ---------------------------------------------------------------------------

class TestApplyColumnWidths:
    def test_sets_column_width(self):
        fmt = _import_formatting()
        wb = Workbook()
        ws = wb.active
        fmt.apply_column_widths(ws, {"A": 20, "B": 35, "C": 12})
        assert ws.column_dimensions["A"].width == 20
        assert ws.column_dimensions["B"].width == 35
        assert ws.column_dimensions["C"].width == 12

    def test_empty_dict_is_noop(self):
        fmt = _import_formatting()
        wb = Workbook()
        ws = wb.active
        # Should not raise
        fmt.apply_column_widths(ws, {})

    def test_does_not_mutate_input_dict(self):
        fmt = _import_formatting()
        wb = Workbook()
        ws = wb.active
        widths = {"A": 15}
        original_widths = dict(widths)
        fmt.apply_column_widths(ws, widths)
        assert widths == original_widths


# ---------------------------------------------------------------------------
# apply_style helper
# ---------------------------------------------------------------------------

class TestApplyStyle:
    def test_applies_font_to_cell(self):
        fmt = _import_formatting()
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        style = fmt.get_wp_style()
        fmt.apply_style(cell, style)
        assert cell.font.bold is True

    def test_applies_fill_to_cell(self):
        fmt = _import_formatting()
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        style = fmt.get_wp_style()
        fmt.apply_style(cell, style)
        assert cell.fill.fgColor.rgb.upper().endswith("D9E2F3")


# ---------------------------------------------------------------------------
# HEADER_FONT constant
# ---------------------------------------------------------------------------

class TestHeaderFont:
    def test_exists(self):
        fmt = _import_formatting()
        assert hasattr(fmt, "HEADER_FONT")

    def test_is_bold(self):
        fmt = _import_formatting()
        assert fmt.HEADER_FONT.bold is True

    def test_size_is_11(self):
        fmt = _import_formatting()
        assert fmt.HEADER_FONT.size == 11
