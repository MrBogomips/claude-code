"""Tests for the WBS sheet generator (helpers/wb_wbs.py).

TDD: these tests are written FIRST and must fail before the implementation exists.
"""
import types

import pytest
from openpyxl import Workbook


# ---------------------------------------------------------------------------
# Helper – build data dict from fixtures
# ---------------------------------------------------------------------------

def _make_data(sample_config, sample_phases, sample_roles):
    return {
        "config": sample_config,
        "phases": sample_phases,
        "roles": sample_roles,
    }


def _get_ws(wb):
    return wb["WBS"]


# ---------------------------------------------------------------------------
# Import the module under test
# ---------------------------------------------------------------------------

from helpers import wb_wbs  # noqa: E402  (import after path setup handled by pytest)


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

class TestBuildCreatesSheet:
    def test_build_creates_sheet(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wb_wbs.build(wb, data)
        assert "WBS" in wb.sheetnames


class TestHeaderRow:
    def _header_values(self, ws):
        return [ws.cell(row=1, column=c).value for c in range(1, 20)]

    def test_header_row(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        wb_wbs.build(wb, data)
        ws = _get_ws(wb)
        headers = self._header_values(ws)
        eu = sample_config["effort_unit"]    # "pd"
        du = sample_config["duration_unit"]  # "d"
        expected = [
            "ID",
            "Phase",
            "Work Package",
            "Activity",
            f"Best Effort ({eu})",
            f"Likely Effort ({eu})",
            f"Worst Effort ({eu})",
            f"PERT Effort ({eu})",
            f"Best Duration ({du})",
            f"Likely Duration ({du})",
            f"Worst Duration ({du})",
            f"PERT Duration ({du})",
            f"\u03c3 Duration",
            "Resources",
            "Dependencies",
            "Risks",
            "Notes",
            "Billable",
            "Billable PERT Effort",
        ]
        assert headers == expected

    def test_header_row_custom_units(self, sample_phases, sample_roles):
        wb = Workbook()
        cfg = {
            "lang": "en",
            "effort_unit": "hours",
            "duration_unit": "weeks",
            "primary_color": "1B4FA5",
            "currency": "EUR",
            "period_type": "biweekly",
            "start_date": "2026-04-06",
            "management_reserve_pct": 0.10,
            "avg_rate": 500,
        }
        data = _make_data(cfg, sample_phases, sample_roles)
        wb_wbs.build(wb, data)
        ws = _get_ws(wb)
        h = self._header_values(ws)
        assert h[4] == "Best Effort (hours)"
        assert h[8] == "Best Duration (weeks)"


class TestPhaseRowFormulas:
    def test_phase_row_formulas(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        # Phase "1" is the first data row (row 2)
        phase_rows = info["phase_rows"]
        assert len(phase_rows) >= 1
        r = phase_rows[0]

        # E,F,G => SUM formulas
        for col in ("E", "F", "G"):
            val = ws[f"{col}{r}"].value
            assert isinstance(val, str) and val.startswith("=SUM("), \
                f"Expected SUM formula in {col}{r}, got {val!r}"

        # I,J,K => SUM formulas
        for col in ("I", "J", "K"):
            val = ws[f"{col}{r}"].value
            assert isinstance(val, str) and val.startswith("=SUM("), \
                f"Expected SUM formula in {col}{r}, got {val!r}"

        # H => PERT effort formula
        h_val = ws[f"H{r}"].value
        assert isinstance(h_val, str) and h_val.startswith("="), \
            f"Expected PERT formula in H{r}, got {h_val!r}"

        # L => PERT duration formula
        l_val = ws[f"L{r}"].value
        assert isinstance(l_val, str) and l_val.startswith("="), \
            f"Expected PERT formula in L{r}, got {l_val!r}"

        # M => σ formula
        m_val = ws[f"M{r}"].value
        assert isinstance(m_val, str) and m_val.startswith("="), \
            f"Expected sigma formula in M{r}, got {m_val!r}"

        # S => SUM for billable
        s_val = ws[f"S{r}"].value
        assert isinstance(s_val, str) and s_val.startswith("=SUM("), \
            f"Expected SUM formula in S{r}, got {s_val!r}"


class TestWpRowFormulas:
    def test_wp_row_formulas(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        # Find WP rows by looking for id "1.1"
        wp_row = None
        for r in range(2, info["total_row"]):
            if ws[f"A{r}"].value == "1.1":
                wp_row = r
                break
        assert wp_row is not None, "WP row with id '1.1' not found"

        # E,F,G => SUM formulas
        for col in ("E", "F", "G"):
            val = ws[f"{col}{wp_row}"].value
            assert isinstance(val, str) and val.startswith("=SUM("), \
                f"Expected SUM formula in {col}{wp_row}, got {val!r}"


class TestLeafRowValues:
    def _find_leaf_row(self, ws, leaf_id, total_row):
        for r in range(2, total_row):
            if ws[f"A{r}"].value == leaf_id:
                return r
        return None

    def test_leaf_row_values(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        r = self._find_leaf_row(ws, "1.1.1", info["total_row"])
        assert r is not None, "Leaf row '1.1.1' not found"

        # E,F,G => numeric values, not formulas
        for col in ("E", "F", "G"):
            val = ws[f"{col}{r}"].value
            assert isinstance(val, (int, float)), \
                f"Expected numeric in {col}{r}, got {val!r}"
            assert not (isinstance(val, str) and val.startswith("=")), \
                f"Got formula in {col}{r}: {val!r}"

        # H => PERT formula
        h_val = ws[f"H{r}"].value
        assert isinstance(h_val, str) and h_val.startswith("="), \
            f"Expected PERT formula in H{r}, got {h_val!r}"

        # R => "Y" or "N"
        r_val = ws[f"R{r}"].value
        assert r_val in ("Y", "N"), f"Expected Y or N in R{r}, got {r_val!r}"

        # S => IF formula
        s_val = ws[f"S{r}"].value
        assert isinstance(s_val, str) and s_val.startswith("=IF("), \
            f"Expected IF formula in S{r}, got {s_val!r}"


class TestTotalRow:
    def test_total_row(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        tr = info["total_row"]
        # Columns E through S should have SUM formulas referencing phase rows
        for col in ("E", "F", "G", "H", "I", "J", "K", "L", "M", "S"):
            val = ws[f"{col}{tr}"].value
            assert isinstance(val, str) and val.startswith("=SUM("), \
                f"Expected SUM formula in {col}{tr}, got {val!r}"


class TestFormattingApplied:
    def test_formatting_applied(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        phase_rows = info["phase_rows"]
        total_row = info["total_row"]

        # Phase rows: bold white font (from get_phase_style)
        for pr in phase_rows:
            cell = ws[f"A{pr}"]
            assert cell.font.bold is True, f"Phase row {pr} A should be bold"
            assert cell.font.color.rgb == "FFFFFFFF", \
                f"Phase row {pr} A should have white font, got {cell.font.color.rgb}"

        # Total row: bold white on dark background
        cell = ws[f"A{total_row}"]
        assert cell.font.bold is True, "Total row A should be bold"
        assert cell.font.color.rgb == "FFFFFFFF", \
            "Total row A should have white font"

        # Formula columns H, L, M, S on a leaf row should have FFF2CC fill
        leaf_row = None
        for r in range(2, total_row):
            if ws[f"A{r}"].value == "1.1.1":
                leaf_row = r
                break
        assert leaf_row is not None
        for col in ("H", "L", "M", "S"):
            cell = ws[f"{col}{leaf_row}"]
            # fgColor may be stored as RGB or theme; check the pattern fill type
            assert cell.fill.fill_type == "solid", \
                f"Formula col {col} on leaf row should have solid fill"
            assert cell.fill.fgColor.rgb == "FFFFF2CC", \
                f"Formula col {col} on leaf row should have FFF2CC fill, got {cell.fill.fgColor.rgb}"


class TestNoFormulaInInputCells:
    def test_no_formula_in_input_cells(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        for r in range(info["data_start_row"], info["total_row"]):
            if r in info["phase_rows"]:
                continue
            # Check if this is a WP row (has id with exactly one dot)
            row_id = ws[f"A{r}"].value
            if row_id and isinstance(row_id, str) and row_id.count(".") == 1:
                continue  # WP row, skip
            # Leaf row: E,F,G must be numeric
            for col in ("E", "F", "G"):
                val = ws[f"{col}{r}"].value
                assert not (isinstance(val, str) and val.startswith("=")), \
                    f"Found formula in input cell {col}{r}: {val!r}"


class TestNoValueInFormulaCells:
    def test_no_value_in_formula_cells(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        # All data rows (including phase, WP, leaf) must have formulas in H, L, M, S
        for r in range(info["data_start_row"], info["total_row"]):
            for col in ("H", "L", "M", "S"):
                val = ws[f"{col}{r}"].value
                assert isinstance(val, str) and val.startswith("="), \
                    f"Expected formula in {col}{r}, got {val!r}"


class TestBillableFlag:
    def test_billable_flag(self, sample_config, sample_phases, sample_roles):
        wb = Workbook()
        data = _make_data(sample_config, sample_phases, sample_roles)
        info = wb_wbs.build(wb, data)
        ws = _get_ws(wb)

        # "2.1.2" has billable: False
        non_billable_row = None
        for r in range(info["data_start_row"], info["total_row"]):
            if ws[f"A{r}"].value == "2.1.2":
                non_billable_row = r
                break
        assert non_billable_row is not None, "Row for activity 2.1.2 not found"

        r_cell_val = ws[f"R{non_billable_row}"].value
        assert r_cell_val == "N", \
            f"Expected 'N' in R{non_billable_row}, got {r_cell_val!r}"

        s_val = ws[f"S{non_billable_row}"].value
        assert isinstance(s_val, str) and s_val.startswith("=IF("), \
            f"Expected IF formula in S{non_billable_row}, got {s_val!r}"
