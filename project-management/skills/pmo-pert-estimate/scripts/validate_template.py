#!/usr/bin/env python3
"""Validate a .xlsx file as a valid pmo-pert-estimate PERT template.

Usage (module):
    from validate_template import validate
    result = validate("path/to/template.xlsx")

Usage (CLI):
    python3 validate_template.py --template path/to/template.xlsx
    # Exits 0 if valid, 1 if invalid.  Prints JSON to stdout.

Return value from validate():
    {
        "valid": True | False,
        "column_map": {
            "WBS": {"ID": "A", "Phase": "B", ...},
            "Resource Plan": {...},
            "Risks": {...},
            "Summary": {...},
        },
        "errors": ["..."],
        "warnings": ["..."],
    }

Sheet names follow the English defaults emitted by the generator. Italian
deliverables are renamed by the localization layer; the validator accepts
both English and Italian sheet names as equivalents.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Required sheets \u2014 canonical English names + Italian aliases
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = ["WBS", "Resource Plan", "Risks", "Summary"]

#: Italian aliases \u2014 accepted in addition to the canonical English names.
SHEET_ALIASES: dict[str, set[str]] = {
    "WBS": {"WBS"},
    "Resource Plan": {"Resource Plan", "Pianificazione Risorse"},
    "Risks": {"Risks", "Rischi"},
    "Summary": {"Summary", "Riepilogo"},
}


# ---------------------------------------------------------------------------
# Required column name fragments per sheet
# (partial matching: a header "Best Effort (pd)" satisfies "Best Effort")
# ---------------------------------------------------------------------------

REQUIRED_COLUMNS: dict[str, list[str]] = {
    "WBS": [
        "ID",
        "Phase",
        "Work Package",
        "Activity",
        "Best Effort",
        "Likely Effort",
        "Worst Effort",
        "PERT Effort",
        "Best Duration",
        "Likely Duration",
        "Worst Duration",
        "PERT Duration",
        "\u03c3 Duration",
        "Resources",
        "Dependencies",
        "Risks",
        "Notes",
        "Billable",
        "Billable PERT Effort",
    ],
    "Resource Plan": [
        # Role / Code / Type, plus at least one week column and the TOTAL.
        # Match accepts both English and Italian via partial matching below.
        "Role",
        "Code",
        "Type",
        "TOTAL",
    ],
    "Risks": [
        "ID",
        "Risk Description",
        "Category",
        "Affected Phases",
        "Probability",
        "Impact",
        "Risk Score",
        "Priority",
        "Strategy",
        "Mitigation Action",
        "Owner",
        "Contingency",
    ],
    "Summary": [
        "Phase",
        "Description",
    ],
}

#: Italian column-name aliases for Resource Plan, used by the partial matcher.
COLUMN_ALIASES: dict[str, dict[str, list[str]]] = {
    "Resource Plan": {
        "Role": ["Role", "Ruolo"],
        "Code": ["Code", "Codice"],
        "Type": ["Type", "Tipo"],
        "TOTAL": ["TOTAL", "TOTALE"],
    },
}

# ---------------------------------------------------------------------------
# PERT formula pattern for WBS column H (PERT Effort)
# Matches:  =(E{n}+4*F{n}+G{n})/6
# Also accepts SUM-based aggregation rows (WP / phase rows).
# ---------------------------------------------------------------------------

_PERT_EFFORT_RE = re.compile(
    r"^\s*=\s*\(\s*E\d+\s*\+\s*4\s*\*\s*F\d+\s*\+\s*G\d+\s*\)\s*/\s*6\s*$",
    re.IGNORECASE,
)
_SUM_RE = re.compile(r"^\s*=\s*SUM\s*\(", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Header row per canonical sheet (1-based)
# ---------------------------------------------------------------------------

_HEADER_ROW: dict[str, int] = {
    "WBS": 1,
    "Resource Plan": 1,
    "Risks": 1,
    "Summary": 1,
}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def validate(template_path: str) -> dict[str, Any]:
    """Validate a .xlsx file as a valid PERT template.

    Args:
        template_path: Path to the .xlsx file to validate.

    Returns:
        A dict with keys ``valid``, ``column_map``, ``errors``, ``warnings``.
    """
    errors: list[str] = []
    warnings: list[str] = []
    column_map: dict[str, dict[str, str]] = {}

    # ------------------------------------------------------------------
    # Open workbook
    # ------------------------------------------------------------------
    path = Path(template_path)
    if not path.exists():
        return _result(False, column_map, [f"File not found: {template_path}"], warnings)

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
    except Exception as exc:
        return _result(False, column_map, [f"Cannot open file: {exc}"], warnings)

    # ------------------------------------------------------------------
    # Resolve canonical sheets — accept English defaults and Italian aliases.
    # ------------------------------------------------------------------
    present_sheets = set(wb.sheetnames)
    sheet_resolution: dict[str, str] = {}  # canonical -> actual name in workbook
    for canonical in REQUIRED_SHEETS:
        aliases = SHEET_ALIASES.get(canonical, {canonical})
        match = next((name for name in aliases if name in present_sheets), None)
        if match is None:
            errors.append(f"Missing sheet: {canonical}")
        else:
            sheet_resolution[canonical] = match

    if errors:
        wb.close()
        return _result(False, column_map, errors, warnings)

    # ------------------------------------------------------------------
    # For each required sheet: scan headers, build column_map
    # ------------------------------------------------------------------
    for canonical in REQUIRED_SHEETS:
        actual = sheet_resolution[canonical]
        ws = wb[actual]
        header_row = _HEADER_ROW.get(canonical, 1)
        sheet_col_map, sheet_errors, sheet_warnings = _validate_sheet_columns(
            ws,
            actual,
            header_row,
            REQUIRED_COLUMNS.get(canonical, []),
            aliases=COLUMN_ALIASES.get(canonical, {}),
        )
        column_map[canonical] = sheet_col_map
        errors.extend(sheet_errors)
        warnings.extend(sheet_warnings)

    # ------------------------------------------------------------------
    # Validate PERT formula pattern in WBS column H
    # ------------------------------------------------------------------
    if "WBS" not in [e.split(":")[1].strip() if ":" in e else "" for e in errors]:
        wbs_errors = _validate_wbs_pert_formulas(wb[sheet_resolution["WBS"]])
        errors.extend(wbs_errors)

    wb.close()

    valid = len(errors) == 0
    return _result(valid, column_map, errors, warnings)


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------

def _result(
    valid: bool,
    column_map: dict,
    errors: list[str],
    warnings: list[str],
) -> dict[str, Any]:
    return {
        "valid": valid,
        "column_map": column_map,
        "errors": errors,
        "warnings": warnings,
    }


def _validate_sheet_columns(
    ws,
    sheet_name: str,
    header_row: int,
    required: list[str],
    aliases: dict[str, list[str]] | None = None,
) -> tuple[dict[str, str], list[str], list[str]]:
    """Scan the header row of *ws* and return (col_map, errors, warnings).

    Matching strategy: each required fragment is matched to at most one header,
    and each header is consumed by at most one requirement. Requirements are
    matched in descending order of specificity (longest fragment first) so that
    "Billable PERT Effort" is satisfied before the shorter "Billable" fragment.

    If a fragment has aliases (e.g. English/Italian translations), any alias
    can satisfy it.

    Args:
        ws:         Worksheet to inspect.
        sheet_name: Sheet name (for error messages).
        header_row: Row number (1-based) containing column headers.
        required:   List of required column name fragments.
        aliases:    Optional ``{canonical_fragment: [alternates...]}``.

    Returns:
        col_map:  {header_text: column_letter} for all headers found.
        errors:   List of error strings for missing required columns.
        warnings: List of warning strings for extra (unexpected) columns.
    """
    errors: list[str] = []
    warnings: list[str] = []
    col_map: dict[str, str] = {}
    aliases = aliases or {}

    row_cells = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]
    for cell in row_cells:
        value = cell.value
        if value is not None and str(value).strip():
            col_letter = get_column_letter(cell.column)
            col_map[str(value).strip()] = col_letter

    available_headers = set(col_map.keys())
    assigned: dict[str, str] = {}

    for req in sorted(required, key=len, reverse=True):
        candidates = aliases.get(req, [req])
        matched: str | None = None
        for candidate in candidates:
            matched = _find_partial_match(candidate, available_headers)
            if matched:
                break
        if matched:
            assigned[req] = matched
            available_headers.discard(matched)

    for req in required:
        if req not in assigned:
            errors.append(f"Missing column '{req}' in sheet '{sheet_name}'")

    for header in available_headers:
        warnings.append(f"Extra column '{header}' in sheet '{sheet_name}'")

    return col_map, errors, warnings


def _find_partial_match(fragment: str, headers: set[str]) -> str | None:
    """Return the first header in *headers* that contains *fragment* (case-insensitive)."""
    fragment_lower = fragment.lower()
    for header in sorted(headers):
        if fragment_lower in header.lower():
            return header
    return None


def _validate_wbs_pert_formulas(ws) -> list[str]:
    """Check that PERT Effort cells in WBS column H use the expected formula.

    Column H (index 8) should contain either:
    - A PERT formula: =(E{n}+4*F{n}+G{n})/6
    - A SUM aggregation formula (for WP / phase rows)
    - A numeric value (for the TOTAL row)
    - None / empty

    Returns a list of error strings (empty if all formulas are valid).
    """
    errors: list[str] = []
    pert_col = 8  # column H

    # Skip header row (row 1)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=pert_col, max_col=pert_col), start=2):
        cell = row[0]
        value = cell.value
        if value is None:
            continue
        if isinstance(value, (int, float)):
            continue  # numeric value is fine (e.g. TOTAL row computed value)
        if isinstance(value, str) and value.startswith("="):
            formula = value.strip()
            if _SUM_RE.match(formula):
                continue  # SUM aggregation is acceptable
            if _PERT_EFFORT_RE.match(formula):
                continue  # correct PERT pattern
            errors.append(
                f"WBS column H row {row_idx}: unexpected formula '{formula}'; "
                f"expected PERT pattern =(E{{n}}+4*F{{n}}+G{{n}})/6 or SUM aggregation"
            )

    return errors


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _cli_main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Validate a .xlsx file as a pmo-pert-estimate PERT template."
    )
    parser.add_argument(
        "--template",
        required=True,
        help="Path to the .xlsx file to validate.",
    )
    args = parser.parse_args(argv)

    result = validate(args.template)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0 if result["valid"] else 1


if __name__ == "__main__":
    sys.exit(_cli_main())
